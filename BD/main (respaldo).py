import streamlit as st
import pandas as pd
import requests
import plotly.graph_objects as go
import os, base64, time, tempfile
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from fpdf import FPDF

# ── PALETA ───────────────────────────────────────────────────────────────────
C = {
    "primary":       "#4d6bff",
    "primary_soft":  "rgba(77,107,255,0.08)",
    "n900":          "#0f172a",
    "n600":          "#475569",
    "n400":          "#94a3b8",
    "n100":          "#f1f5f9",
    "bg":            "#f8faff",
    "card":          "rgba(255,255,255,0.85)",
    "border":        "rgba(77,107,255,0.10)",
}

BRAND_COLORS = {
    "Banco de Chile":          "#0033A0",   # azul oscuro
    "Banco Internacional":     "#005EB8",
    "Banco Estado":            "#F5A800",   # dorado
    "Scotiabank Chile":        "#ED1C24",
    "Bci":                     "#0057A8",
    "Banco Do Brasil":         "#F5A800",
    "JP Morgan Chase":         "#1A1A2E",
    "China Construction Bank": "#CC0000",
    "Santander Chile":         "#EC0000",
    "Itaú Corpbanca":          "#FF6200",
    "Banco Itaú":              "#FF6200",
    "BTG Pactual Chile":       "#003366",
    "Banco Security":          "#005BAA",
    "Banco Falabella":         "#00A651",   # verde
    "Banco BICE":              "#1A3A6C",
    "Banco Consorcio":         "#00457C",
    "Banco Ripley":            "#6B2D8B",
    "Banco ST":                "#2C7BE5",
}

st.set_page_config(page_title="INSAIT Pro", layout="wide", initial_sidebar_state="collapsed")

RUTA_INS_T = r"C:\Users\Hp\Desktop\Tesis\INS T"
RUTA_BD    = r"C:\Users\Hp\Desktop\Tesis\BD"
API_KEY    = "ea9b378f0cbb6fc8b27040141e054e22752373f2"

# ── PDF ───────────────────────────────────────────────────────────────────────
class ReportePDF(FPDF):
    def header(self):
        logo_path = os.path.join(RUTA_INS_T, "1.png")
        if os.path.exists(logo_path): self.image(logo_path, x=10, y=8, w=30)
        self.set_font('Arial','B',10); self.set_text_color(150,150,150)
        self.cell(0,10,f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}',0,1,'R')
        self.ln(10); self.set_font('Arial','B',16); self.set_text_color(35,61,255)
        self.cell(0,10,'REPORTE FINANCIERO',0,1,'C'); self.ln(5)
    def footer(self):
        self.set_y(-15); self.set_font('Arial','I',8); self.set_text_color(128,128,128)
        self.cell(0,10,f'Página {self.page_no()}',0,0,'C')

def generar_pdf_profesional(df_filt, ratios_data, sel_bancos, sel_cuentas):
    pdf = ReportePDF(); pdf.set_auto_page_break(auto=True, margin=15); pdf.add_page()
    pdf.set_fill_color(35,61,255); pdf.set_text_color(255,255,255); pdf.set_font("Arial",'B',11)
    pdf.cell(0,10," 1. ANALISIS DE CUENTAS POR PERIODO (DATOS NOMINALES)",0,1,'L',True); pdf.ln(5)
    for banco in sel_bancos:
        df_banco = df_filt[df_filt["Banco"]==banco].sort_values(["Anho","Cuenta"])
        if df_banco.empty: continue
        pdf.set_text_color(35,61,255); pdf.set_font("Arial",'B',10)
        pdf.cell(0,8,f">> INSTITUCION: {banco}",0,1)
        pdf.set_fill_color(240,240,240); pdf.set_text_color(0,0,0); pdf.set_font("Arial",'B',9)
        pdf.cell(40,8,"Periodo",1,0,'C',True); pdf.cell(100,8,"Cuenta",1,0,'C',True)
        pdf.cell(50,8,"Valor (MM$)",1,1,'C',True); pdf.set_font("Arial",'',9)
        for _,row in df_banco.iterrows():
            pdf.cell(40,7,str(row['Anho']),1,0,'C'); pdf.cell(100,7,str(row['Cuenta']),1,0,'L')
            pdf.cell(50,7,f"{row['Valor']:,.0f}",1,1,'R')
        pdf.ln(5)
    pdf.add_page(); pdf.set_fill_color(35,61,255); pdf.set_text_color(255,255,255)
    pdf.set_font("Arial",'B',11)
    pdf.cell(0,10," 2. VISUALIZACION DE TENDENCIAS GRAFICAS",0,1,'L',True); pdf.ln(5)
    combos = [(b,ct) for b in sel_bancos for ct in sel_cuentas
              if not df_filt[(df_filt["Banco"]==b)&(df_filt["Cuenta"]==ct)].empty]
    with tempfile.TemporaryDirectory() as tmpdir:
        for idx,(b,ct) in enumerate(combos):
            d_p = df_filt[(df_filt["Banco"]==b)&(df_filt["Cuenta"]==ct)].sort_values("Anho")
            fig_p = go.Figure()
            fig_p.add_trace(go.Scatter(x=d_p["Anho"],y=d_p["Valor"],fill='tozeroy',
                                       fillcolor='rgba(35,61,255,0.1)',
                                       line=dict(color=C["primary"],width=4),mode='lines+markers'))
            fig_p.update_layout(title=f"{b}: {ct}",template="plotly_white",width=700,height=400)
            img_p = os.path.join(tmpdir,f"c_{idx}.png"); fig_p.write_image(img_p,scale=2)
            if pdf.get_y()>200: pdf.add_page()
            pdf.image(img_p,x=20,w=170); pdf.ln(2)
    return pdf.output(dest='S').encode('latin-1')

# ── DATOS ─────────────────────────────────────────────────────────────────────
def b64(p):
    try:
        if os.path.exists(p):
            with open(p,"rb") as f: return base64.b64encode(f.read()).decode()
    except: pass
    return None

def _fetch(c, a):
    u = f"https://api.cmfchile.cl/api-sbifv3/recursos_api/balances/{a}/12/instituciones/{c}?apikey={API_KEY}&formato=json"
    try:
        r = requests.get(u,timeout=12).json()
        return [{"Banco":c,"Anho":int(a),"Cuenta":i["DescripcionCuenta"].strip().upper(),
                 "Valor":float(i["MonedaTotal"].replace(",","."))/1e6}
                for i in r.get("CodigosBalances",[])]
    except: return []

@st.cache_data(show_spinner=False)
def cargar_bancos():
    m = {"001":"Banco de Chile","009":"Banco Internacional","012":"Banco Estado",
         "014":"Scotiabank Chile","016":"Bci","027":"Banco Do Brasil","028":"JP Morgan Chase",
         "031":"China Construction Bank","037":"Santander Chile","039":"Itaú Corpbanca",
         "040":"Banco Itaú","041":"BTG Pactual Chile","049":"Banco Security",
         "051":"Banco Falabella","053":"Banco BICE","055":"Banco Consorcio",
         "057":"Banco Ripley","067":"Banco ST"}
    rows = []
    with ThreadPoolExecutor(max_workers=15) as ex:
        futs = [ex.submit(_fetch,c,a) for c in m.keys() for a in range(2021,2026)]
        for f in as_completed(futs): rows.extend(f.result())
    df = pd.DataFrame(rows)
    if not df.empty: df["Banco"] = df["Banco"].map(m)
    return df

# ── SESSION STATE ─────────────────────────────────────────────────────────────
if "cargado" not in st.session_state:
    ph = st.empty()
    with ph.container():
        st.markdown(f'<div class="minimal-load"><div class="spinner"></div>'
                    f'<div style="font-weight:600;color:{C["primary"]};">INSAIT Pro</div></div>',
                    unsafe_allow_html=True)
    st.session_state["df_b"] = cargar_bancos()
    st.session_state["cargado"] = True
    time.sleep(1); ph.empty()

for k,v in [("all_b",False),("all_a",False),("all_c",False),("all_r",False)]:
    if k not in st.session_state: st.session_state[k] = v

df_b        = st.session_state["df_b"]
all_banks_s = sorted(df_b["Banco"].unique()) if not df_b.empty else []
BANK_COLOR  = {b: BRAND_COLORS.get(b, "#233dff") for b in all_banks_s}

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ══════════════════════════════════════════════════════════════════
   BASE & FONDO
   ══════════════════════════════════════════════════════════════════ */
html, body, .stApp {{
    background-color: {C["bg"]} !important;
    font-family: 'Inter', sans-serif !important;
    color: {C["n900"]} !important;
}}
/* Gradiente azul leve en fondo */
.stApp::before {{
    content: '';
    position: fixed;
    top: -20%;
    left: -10%;
    width: 60%;
    height: 60%;
    background: radial-gradient(ellipse, rgba(77,107,255,0.09) 0%, transparent 70%);
    pointer-events: none;
    z-index: 0;
}}
.stApp::after {{
    content: '';
    position: fixed;
    bottom: -20%;
    right: -5%;
    width: 50%;
    height: 50%;
    background: radial-gradient(ellipse, rgba(77,107,255,0.06) 0%, transparent 70%);
    pointer-events: none;
    z-index: 0;
}}
label, p, div, span, .stMarkdown {{
    font-family: 'Inter', sans-serif !important;
    color: {C["n900"]};
}}
.main .block-container {{
    max-width: 780px !important;
    padding-left: 2rem !important;
    padding-right: 2rem !important;
    padding-top: 1.5rem !important;
    padding-bottom: 60px !important;
    margin-left: auto !important;
    margin-right: auto !important;
    position: relative;
    z-index: 1;
}}

/* ══════════════════════════════════════════════════════════════════
   OVERRIDES BASEWEB / STREAMLIT
   ══════════════════════════════════════════════════════════════════ */
:root {{ --primary-color: {C["primary"]} !important; }}

/* Toggle activo */
[data-baseweb="checkbox"] [data-checked="true"],
[data-baseweb="toggle"] [data-checked="true"],
[data-testid="stToggle"] [aria-checked="true"],
[role="switch"][aria-checked="true"] {{
    background-color: {C["primary"]} !important;
    border-color: {C["primary"]} !important;
}}
div[data-testid="stToggle"] label span {{
    font-size: 13px !important;
    color: {C["n600"]} !important;
    font-family: 'Inter', sans-serif !important;
}}

/* Tab highlight/border */
[data-baseweb="tab-highlight"] {{ background-color: {C["primary"]} !important; }}
[data-baseweb="tab-border"],
div[data-testid="stTabBar"] > div > div[style*="height: 3px"],
div[data-testid="stTabBar"] > div > div[style*="height:3px"] {{
    display: none !important;
}}

/* Multiselect chips */
span[data-baseweb="tag"] {{
    background-color: {C["primary_soft"]} !important;
    border-color: rgba(77,107,255,0.2) !important;
}}
span[data-baseweb="tag"] span {{ color: {C["primary"]} !important; }}
span[data-baseweb="tag"] svg {{ fill: {C["primary"]} !important; }}

/* Dropdown */
[data-baseweb="menu"] li:hover,
[data-baseweb="menu"] [aria-selected="true"] {{
    background-color: {C["primary_soft"]} !important;
    color: {C["primary"]} !important;
}}

/* Multiselect input */
div[data-baseweb="select"] > div {{
    background: white !important;
    border: 1px solid {C["border"]} !important;
    border-radius: 12px !important;
}}

/* ══════════════════════════════════════════════════════════════════
   LOGO
   ══════════════════════════════════════════════════════════════════ */
.logo-container {{
    display: flex;
    justify-content: center;
    padding: 36px 0 24px 0;
    width: 100%;
}}
.logo-img {{
    height: 72px;
    filter: drop-shadow(0px 4px 16px rgba(77,107,255,0.15));
}}

/* ── Filtros container ───────────────────────────────────────────── */
div[data-testid="stVerticalBlockBorderWrapper"] {{
    border: 1px solid rgba(77,107,255,0.12) !important;
    border-radius: 20px !important;
    background: rgba(255,255,255,0.85) !important;
    backdrop-filter: blur(20px) !important;
    box-shadow: 0 2px 16px rgba(77,107,255,0.06) !important;
    padding: 8px 4px !important;
}}

/* Botones pequeños para exportar */
div[data-testid="stColumns"] .stButton > button {{
    background: transparent !important;
    color: {C["primary"]} !important;
    border: 1px solid rgba(77,107,255,0.35) !important;
    border-radius: 8px !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    padding: 5px 14px !important;
    box-shadow: none !important;
    letter-spacing: 0.2px !important;
}}
div[data-testid="stColumns"] .stButton > button:hover {{
    background: {C["primary_soft"]} !important;
    border-color: {C["primary"]} !important;
    transform: none !important;
    box-shadow: none !important;
}}
div[data-testid="stColumns"] .stDownloadButton > button {{
    background: transparent !important;
    color: {C["primary"]} !important;
    border: 1px solid rgba(77,107,255,0.35) !important;
    border-radius: 8px !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    padding: 5px 14px !important;
    box-shadow: none !important;
}}

/* ══════════════════════════════════════════════════════════════════
   BOTONES PRINCIPALES
   ══════════════════════════════════════════════════════════════════ */
.stButton > button,
.stDownloadButton > button {{
    background: linear-gradient(135deg, #4d6bff 0%, #3350e8 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    padding: 10px 20px !important;
    box-shadow: 0 4px 14px rgba(77,107,255,0.25) !important;
    transition: all 0.2s !important;
    letter-spacing: 0.2px !important;
}}
.stButton > button:hover,
.stDownloadButton > button:hover {{
    box-shadow: 0 6px 20px rgba(77,107,255,0.38) !important;
    transform: translateY(-1px) !important;
}}

/* ══════════════════════════════════════════════════════════════════
   TABS
   ══════════════════════════════════════════════════════════════════ */
div[data-testid="stTabs"] {{
    background: transparent !important;
    box-shadow: none !important;
    padding: 0 !important;
}}
div[data-testid="stTabBar"] {{
    background: white !important;
    border: 1px solid rgba(77,107,255,0.10) !important;
    border-radius: 14px !important;
    padding: 4px !important;
    margin: 0 auto 20px auto !important;
    overflow: hidden !important;
    display: flex !important;
    justify-content: center !important;
    width: fit-content !important;
    max-width: 100% !important;
    gap: 2px !important;
    box-shadow: 0 2px 12px rgba(77,107,255,0.08) !important;
}}
div[data-testid="stTabBar"] [data-baseweb="tab-highlight"],
div[data-testid="stTabBar"] [data-baseweb="tab-border"] {{
    display: none !important;
}}
div[data-testid="stTabBar"] button[data-baseweb="tab"] {{
    background: transparent !important;
    border: none !important;
    outline: none !important;
    box-shadow: none !important;
    border-radius: 10px !important;
    padding: 8px 18px !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 12.5px !important;
    font-weight: 500 !important;
    color: {C["n400"]} !important;
    white-space: nowrap !important;
    transition: all 0.15s ease !important;
    flex-shrink: 0 !important;
}}
div[data-testid="stTabBar"] button[data-baseweb="tab"] p {{
    color: inherit !important;
    font-weight: inherit !important;
    font-size: 12.5px !important;
    margin: 0 !important;
}}
div[data-testid="stTabBar"] button[aria-selected="true"] {{
    background: {C["primary_soft"]} !important;
    color: {C["primary"]} !important;
    font-weight: 700 !important;
    border: 1px solid rgba(77,107,255,0.15) !important;
}}
div[data-testid="stTabBar"] button[aria-selected="true"] p {{
    color: {C["primary"]} !important;
    font-weight: 700 !important;
}}
div[data-testid="stTabs"] [role="tabpanel"] {{
    padding-top: 4px !important;
    background: transparent !important;
}}

/* ══════════════════════════════════════════════════════════════════
   MÓDULOS / CARDS
   ══════════════════════════════════════════════════════════════════ */
.modulo {{
    background: rgba(255,255,255,0.85);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid rgba(77,107,255,0.08);
    border-radius: 20px;
    padding: 22px 22px 18px 22px;
    margin-bottom: 14px;
    box-shadow: 0 2px 16px rgba(77,107,255,0.06), 0 1px 4px rgba(0,0,0,0.03);
}}
.modulo-titulo {{
    font-size: 10px;
    font-weight: 700;
    color: {C["primary"]};
    text-transform: uppercase;
    letter-spacing: 1.5px;
    margin-bottom: 14px;
    display: block;
    opacity: 0.8;
}}
.section-label {{
    font-size: 12px;
    font-weight: 500;
    color: {C["n600"]};
    margin-bottom: 4px;
    display: block;
}}
.divider-min {{ height:1px; background:{C["border"]}; margin:24px 0; border:none; }}

/* ══════════════════════════════════════════════════════════════════
   RATIO CARDS
   ══════════════════════════════════════════════════════════════════ */
.ratio-row {{ display:flex; gap:12px; overflow-x:auto; padding:4px 2px 20px; flex-wrap:nowrap; -webkit-overflow-scrolling:touch; }}
.ratio-card {{
    min-width:200px; background:white; border-radius:16px; padding:18px 16px;
    border:1px solid rgba(77,107,255,0.08); display:flex; flex-direction:column; gap:5px;
    box-shadow: 0 2px 12px rgba(77,107,255,0.06);
}}
.ratio-card.leader {{
    border: 1.5px solid rgba(77,107,255,0.35);
    background: rgba(77,107,255,0.05);
    box-shadow: 0 4px 16px rgba(77,107,255,0.12);
}}
.ratio-meta {{ font-size:9px; font-weight:700; color:{C["n400"]}; text-transform:uppercase; letter-spacing:1.2px; }}
.ratio-banco {{ font-size:12px; font-weight:600; color:{C["n900"]}; }}
.ratio-divider {{ height:1px; background:{C["n100"]}; margin:6px 0; }}
.ratio-num {{ font-size:28px; font-weight:800; letter-spacing:-1px; line-height:1.1; }}
.badge-lider {{
    background: linear-gradient(135deg, {C["primary"]} 0%, #6b84ff 100%);
    color:white; font-size:8px; padding:2px 8px; border-radius:99px;
    font-weight:700; display:inline-flex; align-items:center; letter-spacing:0.5px;
    box-shadow: 0 2px 8px rgba(77,107,255,0.3);
}}

/* ══════════════════════════════════════════════════════════════════
   TOOLTIP
   ══════════════════════════════════════════════════════════════════ */
.ratio-title-row {{ display:flex; align-items:center; gap:8px; margin:8px 0 10px 0; }}
.ratio-title-text {{ font-weight:600; color:{C["n900"]}; font-size:14px; letter-spacing:-0.2px; }}
.info-popup-wrap {{ position:relative; display:inline-flex; align-items:center; }}
.info-btn {{
    display:inline-flex; align-items:center; justify-content:center;
    width:17px; height:17px; border-radius:50%;
    background:{C["n100"]}; color:{C["n400"]};
    font-size:10px; font-weight:700; cursor:default; user-select:none;
    flex-shrink:0; transition:all 0.15s; border: 1px solid rgba(77,107,255,0.10);
}}
.info-btn:hover {{ background:{C["primary_soft"]}; color:{C["primary"]}; }}
.info-popup {{
    display:none; position:absolute; left:24px; top:50%; transform:translateY(-50%);
    z-index:9999; background:white; border:1px solid rgba(77,107,255,0.15);
    border-radius:16px; padding:16px 18px; width:270px;
    box-shadow: 0 12px 40px rgba(77,107,255,0.12), 0 2px 8px rgba(0,0,0,0.06);
    pointer-events:none;
}}
.info-popup-wrap:hover .info-popup {{ display:block; }}
.info-popup-title {{ font-size:12px; font-weight:700; color:{C["primary"]}; margin-bottom:6px; }}
.info-popup-desc {{ font-size:11px; color:{C["n600"]}; line-height:1.6; margin-bottom:10px; }}
.info-popup-example {{
    background:{C["primary_soft"]}; border:1px solid rgba(77,107,255,0.15);
    border-radius:8px; padding:8px 10px;
    font-size:10px; color:{C["primary"]}; font-weight:500; line-height:1.5;
}}

/* ══════════════════════════════════════════════════════════════════
   CARGA
   ══════════════════════════════════════════════════════════════════ */
.minimal-load {{
    position:fixed; inset:0; background:rgba(248,250,255,0.92);
    backdrop-filter: blur(12px); z-index:99999;
    display:flex; flex-direction:column; align-items:center; justify-content:center;
}}
.spinner {{
    width:28px; height:28px;
    border:2px solid rgba(77,107,255,0.15);
    border-top:2px solid {C["primary"]};
    border-radius:50%; animation:spin 0.85s linear infinite; margin-bottom:14px;
}}
@keyframes spin {{ to {{ transform:rotate(360deg); }} }}
</style>
""", unsafe_allow_html=True)

# ── JS: forzar colores azul en toggle y tabs (infalible) ──────────────────────
st.markdown("""
<script>
(function applyBlue() {
    const PRIMARY = '#4d6bff';
    const SOFT    = 'rgba(77,107,255,0.08)';

    function paint() {
        // ── Toggle tracks activos ──────────────────────────────────────
        document.querySelectorAll(
            '[data-baseweb="checkbox"] [data-checked="true"], ' +
            '[role="switch"][aria-checked="true"], ' +
            '[data-testid="stToggle"] [aria-checked="true"]'
        ).forEach(el => {
            el.style.setProperty('background-color', PRIMARY, 'important');
            el.style.setProperty('border-color', PRIMARY, 'important');
        });

        // ── Tab highlight (línea de color) ─────────────────────────────
        document.querySelectorAll('[data-baseweb="tab-highlight"]').forEach(el => {
            el.style.setProperty('background-color', PRIMARY, 'important');
        });
        document.querySelectorAll('[data-baseweb="tab-border"]').forEach(el => {
            el.style.setProperty('display', 'none', 'important');
        });

        // ── Tab activo: fondo azul suave, texto azul ───────────────────
        document.querySelectorAll('[data-baseweb="tab"][aria-selected="true"]').forEach(el => {
            el.style.setProperty('background', SOFT, 'important');
            el.style.setProperty('color', PRIMARY, 'important');
        });
    }

    // Ejecutar ahora y observar cambios del DOM
    paint();
    const obs = new MutationObserver(paint);
    obs.observe(document.body, { subtree: true, childList: true, attributes: true,
                                  attributeFilter: ['aria-checked','aria-selected','data-checked'] });
})();
</script>
""", unsafe_allow_html=True)

# ── LOGO ──────────────────────────────────────────────────────────────────────
logo_raw = b64(os.path.join(RUTA_INS_T,"1.png"))
if logo_raw:
    st.markdown(f'<div class="logo-container"><img src="data:image/png;base64,{logo_raw}" class="logo-img"></div>',
                unsafe_allow_html=True)

# ── TABS IN-APP ───────────────────────────────────────────────────────────────
tab_inicio, tab_banco, tab_empresa, tab_mi_empresa, tab_config = st.tabs([
    "Inicio", "Banco", "Empresa", "Mi Empresa", "Configuracion"
])

# ══════════════════════════════════════════════════════
#  INICIO
# ══════════════════════════════════════════════════════
with tab_inicio:
    st.markdown(f"""
    <div style="text-align:center;padding:48px 0 32px 0;">
        <div style="font-size:28px;font-weight:700;color:{C['n900']};letter-spacing:-0.5px;margin-bottom:10px;">
            Bienvenido a INSAIT Pro
        </div>
        <div style="font-size:15px;color:{C['n600']};max-width:420px;margin:0 auto;line-height:1.6;">
            Plataforma de análisis financiero bancario y empresarial.
            Selecciona una sección para comenzar.
        </div>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
#  BANCO
# ══════════════════════════════════════════════════════
with tab_banco:
    b_l = sorted(df_b["Banco"].unique())

    # ── FILTROS ───────────────────────────────────────────────────────────────
    with st.container(border=True):
        col1, col2 = st.columns([0.75, 0.25])
        with col1: st.markdown('<span class="section-label">Bancos</span>', unsafe_allow_html=True)
        with col2:
            tog_b = st.toggle("Todos", value=st.session_state["all_b"], key="tog_b")
        if tog_b != st.session_state["all_b"]:
            st.session_state["all_b"] = tog_b; st.rerun()
        sel_bancos = st.multiselect("Bancos", b_l,
                                    default=b_l if st.session_state["all_b"] else [],
                                    label_visibility="collapsed")

        a_l = sorted(df_b["Anho"].unique(), reverse=True)
        col3, col4 = st.columns([0.75, 0.25])
        with col3: st.markdown('<span class="section-label">Periodos</span>', unsafe_allow_html=True)
        with col4:
            tog_a = st.toggle("Todos", value=st.session_state["all_a"], key="tog_a")
        if tog_a != st.session_state["all_a"]:
            st.session_state["all_a"] = tog_a; st.rerun()
        sel_años = st.multiselect("Periodos", a_l,
                                  default=a_l if st.session_state["all_a"] else [],
                                  label_visibility="collapsed")

        c_std = [
            "TOTAL ACTIVOS","TOTAL PASIVOS","PATRIMONIO",
            "UTILIDAD (PÉRDIDA) DEL EJERCICIO","COLOCACIONES NETAS",
            "DEPOSITOS Y OTRAS CAPTACIONES","GASTOS OPERACIONALES",
            "INGRESOS OPERACIONALES","COLOCACIONES VENCIDAS","PROVISIONES",
            "INGRESOS FINANCIEROS NETOS",
        ]
        col5, col6 = st.columns([0.75, 0.25])
        with col5: st.markdown('<span class="section-label">Cuentas</span>', unsafe_allow_html=True)
        with col6:
            tog_c = st.toggle("Todos", value=st.session_state["all_c"], key="tog_c")
        if tog_c != st.session_state["all_c"]:
            st.session_state["all_c"] = tog_c; st.rerun()
        sel_cuentas = st.multiselect("Cuentas", c_std,
                                     default=c_std if st.session_state["all_c"] else [],
                                     label_visibility="collapsed")

    df_filt = df_b[
        (df_b["Banco"].isin(sel_bancos)) &
        (df_b["Cuenta"].isin(sel_cuentas)) &
        (df_b["Anho"].isin(sel_años))
    ]

    # ── RESUMEN + CALIDAD ─────────────────────────────────────────────────────
    if sel_bancos or sel_años or sel_cuentas:
        total_esperado   = len(sel_bancos) * len(sel_años) * len(sel_cuentas) if (sel_bancos and sel_años and sel_cuentas) else 0
        total_real       = len(df_filt[["Banco","Anho","Cuenta"]].drop_duplicates()) if not df_filt.empty else 0
        pct_cobertura    = int(total_real / total_esperado * 100) if total_esperado > 0 else 0
        bancos_sin_datos = [b for b in sel_bancos if df_filt[df_filt["Banco"]==b].empty] if sel_bancos else []
        total_registros  = len(df_filt) if not df_filt.empty else 0

        if pct_cobertura >= 90:   calidad_color, calidad_icon, calidad_txt = "#06d6a0","✓",f"Completos ({pct_cobertura}%)"
        elif pct_cobertura >= 60: calidad_color, calidad_icon, calidad_txt = "#F5A800","⚠",f"Parciales ({pct_cobertura}%)"
        else:                     calidad_color, calidad_icon, calidad_txt = "#e63946","✗",f"Insuficientes ({pct_cobertura}%)"
        aviso_sin = f" · Sin datos: {', '.join(bancos_sin_datos)}" if bancos_sin_datos else ""

        st.markdown(
            f'<div class="modulo" style="padding:14px 20px;">'
            f'<div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;">'
            f'<div style="display:flex;gap:28px;align-items:center;">'
            # Bancos
            f'<div style="text-align:center;">'
            f'<div style="font-size:22px;font-weight:700;color:{C["n900"]};line-height:1;">{len(sel_bancos)}</div>'
            f'<div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Bancos</div></div>'
            # Separador
            f'<div style="width:1px;height:28px;background:{C["n100"]};"></div>'
            # Periodos
            f'<div style="text-align:center;">'
            f'<div style="font-size:22px;font-weight:700;color:{C["n900"]};line-height:1;">{len(sel_años)}</div>'
            f'<div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Periodos</div></div>'
            # Separador
            f'<div style="width:1px;height:28px;background:{C["n100"]};"></div>'
            # Cuentas
            f'<div style="text-align:center;">'
            f'<div style="font-size:22px;font-weight:700;color:{C["n900"]};line-height:1;">{len(sel_cuentas)}</div>'
            f'<div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Cuentas</div></div>'
            # Separador
            f'<div style="width:1px;height:28px;background:{C["n100"]};"></div>'
            # Total registros
            f'<div style="text-align:center;">'
            f'<div style="font-size:22px;font-weight:700;color:{C["primary"]};line-height:1;">{total_registros:,}</div>'
            f'<div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Registros</div></div>'
            f'</div>'
            # Badge calidad
            f'<div style="display:inline-flex;align-items:center;gap:5px;background:{calidad_color}18;'
            f'border:1px solid {calidad_color}44;border-radius:8px;padding:5px 10px;">'
            f'<span style="color:{calidad_color};font-weight:700;font-size:12px;">{calidad_icon}</span>'
            f'<span style="color:{calidad_color};font-size:11px;font-weight:600;">{calidad_txt}{aviso_sin}</span>'
            f'</div></div></div>',
            unsafe_allow_html=True)

    # ── GRÁFICO ───────────────────────────────────────────────────────────────
    st.markdown(
        f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;'
        f'letter-spacing:1.2px;margin:20px 0 10px 0;">Gráfico</div>',
        unsafe_allow_html=True)

    vista_lineas = st.toggle("Líneas", value=False, key="vista_toggle")
    fig = go.Figure()
    v_t = "Líneas" if vista_lineas else "Barras"

    if not df_filt.empty and sel_cuentas:
        for b in sel_bancos:
            for cuenta in sel_cuentas:
                d = df_filt[(df_filt["Banco"]==b)&(df_filt["Cuenta"]==cuenta)].sort_values("Anho")
                if not d.empty:
                    bc  = BANK_COLOR.get(b, C["primary"])
                    lbl = f"{b} — {cuenta.title()}" if len(sel_cuentas) > 1 else b
                    if v_t == "Barras":
                        fig.add_trace(go.Bar(x=d["Anho"], y=d["Valor"], name=lbl,
                                             marker_color=bc, marker_line_width=0))
                    else:
                        fig.add_trace(go.Scatter(x=d["Anho"], y=d["Valor"], name=lbl,
                                                 mode='lines+markers',
                                                 line=dict(color=bc, width=2.5),
                                                 marker=dict(color=bc, size=7)))
        tit = ", ".join([c.title() for c in sel_cuentas])
    else:
        fig.add_trace(go.Scatter(
            x=[2021,2022,2023,2024,2025], y=[30,45,38,58,52],
            fill="tozeroy", fillcolor="rgba(35,61,255,0.07)",
            line=dict(color="rgba(35,61,255,0.22)", width=2, dash="dot"),
            mode="lines", hoverinfo="skip", showlegend=False))
        fig.add_annotation(text="<b>Seleccione datos para visualizar el análisis</b>",
                           xref="paper", yref="paper", x=0.5, y=0.62, showarrow=False,
                           font=dict(size=15, color="rgba(35,61,255,0.50)", family="DM Sans"))
        fig.add_annotation(text="Elija un banco, periodo y cuenta para comenzar",
                           xref="paper", yref="paper", x=0.5, y=0.50, showarrow=False,
                           font=dict(size=12, color="rgba(75,85,99,0.45)", family="DM Sans"))
        tit = "Vista Previa"

    fig.update_layout(
        title=dict(text=tit, font=dict(size=13, color=C["n600"], family="DM Sans")),
        template="plotly_white", height=320,
        margin=dict(l=0, r=0, t=40, b=10),
        legend=dict(orientation="h", y=-0.35, font=dict(family="DM Sans", size=11)),
        barmode="group",
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── RANKING ───────────────────────────────────────────────────────────────
    if sel_bancos and sel_años and sel_cuentas and not df_filt.empty:
        st.markdown(
            f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;'
            f'letter-spacing:1.2px;margin:20px 0 10px 0;">Ranking</div>',
            unsafe_allow_html=True)
        cuenta_rank = st.selectbox("Cuenta para ranking", sel_cuentas,
                                   key="cuenta_rank_sel", label_visibility="collapsed")
        a_rank  = sel_años[0]
        df_rank = df_filt[(df_filt["Cuenta"]==cuenta_rank)&(df_filt["Anho"]==a_rank)]\
                    .sort_values("Valor", ascending=False).reset_index(drop=True)
        if not df_rank.empty:
            st.markdown(f'<div style="font-size:12px;color:{C["n400"]};margin-bottom:12px;">'
                        f'{cuenta_rank.title()} · {a_rank}</div>', unsafe_allow_html=True)
            medals  = ["🥇","🥈","🥉"]
            max_val = df_rank["Valor"].max() or 1
            html_rank = '<div style="display:flex;flex-direction:column;gap:7px;margin-bottom:4px;">'
            for i, row in df_rank.iterrows():
                pct = row["Valor"] / max_val * 100 if max_val > 0 else 0
                med = medals[i] if i < 3 else f'<span style="font-size:11px;color:{C["n400"]};min-width:22px;display:inline-block;text-align:center;">#{i+1}</span>'
                bc  = BANK_COLOR.get(row["Banco"], C["primary"])
                html_rank += (
                    f'<div style="display:flex;align-items:center;gap:10px;">'
                    f'<span style="font-size:15px;min-width:22px;text-align:center;">{med}</span>'
                    f'<div style="flex:1;">'
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:3px;">'
                    f'<span style="font-size:12px;font-weight:600;color:{C["n900"]};">{row["Banco"]}</span>'
                    f'<span style="font-size:12px;font-weight:700;color:{bc};">{row["Valor"]:,.0f} MM$</span>'
                    f'</div>'
                    f'<div style="background:{C["n100"]};border-radius:99px;height:5px;">'
                    f'<div style="background:{bc};width:{pct:.1f}%;height:5px;border-radius:99px;"></div>'
                    f'</div></div></div>')
            st.markdown(html_rank + '</div>', unsafe_allow_html=True)

    # ── INDICADORES ───────────────────────────────────────────────────────────
    if sel_bancos:
        st.markdown(
            f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;'
            f'letter-spacing:1.2px;margin:20px 0 10px 0;">Indicadores Financieros</div>',
            unsafe_allow_html=True)

        M_CONF = {
            # ── Financieros ──────────────────────────────────────────────
            "Liquidez Corriente": {
                "f":"ACTIVO / PASIVO","n1":"TOTAL ACTIVOS","n2":"TOTAL PASIVOS",
                "grupo":"Financieros",
                "desc":"Mide cuántos activos tiene el banco por cada peso de deuda. Valores mayores a 1 indican solidez.",
                "ejemplo":"Si el banco tiene $200 en activos y $100 en pasivos → Liquidez = 2.0 (cubre 2 veces sus deudas).",
            },
            "Solvencia": {
                "f":"PATRIMONIO / PASIVO","n1":"PATRIMONIO","n2":"TOTAL PASIVOS",
                "grupo":"Financieros",
                "desc":"Indica cuánto patrimonio propio respalda las deudas del banco.",
                "ejemplo":"Patrimonio $50, Pasivos $200 → Solvencia = 0.25.",
            },
            "ROA": {
                "f":"UTILIDAD / ACTIVOS","n1":"UTILIDAD (PÉRDIDA) DEL EJERCICIO","n2":"TOTAL ACTIVOS",
                "grupo":"Financieros",
                "desc":"Mide la rentabilidad del banco en relación a sus activos totales.",
                "ejemplo":"Utilidad $10, Activos $500 → ROA = 0.02 (2%).",
            },
            "ROE": {
                "f":"UTILIDAD / PATRIMONIO","n1":"UTILIDAD (PÉRDIDA) DEL EJERCICIO","n2":"PATRIMONIO",
                "grupo":"Financieros",
                "desc":"Rentabilidad sobre el capital de los accionistas.",
                "ejemplo":"Utilidad $10, Patrimonio $80 → ROE = 0.125 (12.5%).",
            },
            "NIM": {
                "f":"ING. FINANCIEROS NETOS / ACTIVOS","n1":"INGRESOS FINANCIEROS NETOS","n2":"TOTAL ACTIVOS",
                "grupo":"Financieros",
                "desc":"Margen Financiero Neto. Mide la rentabilidad de las operaciones de intermediación financiera respecto a los activos.",
                "ejemplo":"Ingresos financieros netos $15, Activos $500 → NIM = 0.03 (3% de margen neto sobre activos).",
            },
            "Ratio de Eficiencia": {
                "f":"GASTOS / INGRESOS","n1":"GASTOS OPERACIONALES","n2":"INGRESOS OPERACIONALES",
                "grupo":"Financieros",
                "desc":"Cuánto gasta el banco por cada peso que ingresa. Menor = más eficiente.",
                "ejemplo":"Gastos $60, Ingresos $100 → Eficiencia = 0.60.",
            },
            # ── Calidad de Activos y Riesgos ─────────────────────────────
            "Índice de Morosidad": {
                "f":"COL. VENCIDAS / COL. TOTAL","n1":"COLOCACIONES VENCIDAS","n2":"COLOCACIONES NETAS",
                "grupo":"Calidad de Activos y Riesgos",
                "desc":"Porcentaje de créditos que no se están pagando a tiempo.",
                "ejemplo":"Vencidos $5, Total $100 → Morosidad = 0.05 (5% impagos).",
            },
            "Índice de Riesgo": {
                "f":"PROVISIONES / COLOCACIONES","n1":"PROVISIONES","n2":"COLOCACIONES NETAS",
                "grupo":"Calidad de Activos y Riesgos",
                "desc":"Cuánto ha reservado el banco para cubrir posibles pérdidas.",
                "ejemplo":"Provisiones $8, Colocaciones $100 → Riesgo = 0.08.",
            },
            "Ratio de Capital": {
                "f":"PASIVO / PATRIMONIO","n1":"TOTAL PASIVOS","n2":"PATRIMONIO",
                "grupo":"Calidad de Activos y Riesgos",
                "desc":"Nivel de endeudamiento relativo al capital propio.",
                "ejemplo":"Pasivos $300, Patrimonio $100 → Ratio = 3.0.",
            },
            # ── Gestión Clientes ──────────────────────────────────────────
            "Participación de Mercado": {
                "f":"COL. BANCO / DEPÓSITOS","n1":"COLOCACIONES NETAS","n2":"DEPOSITOS Y OTRAS CAPTACIONES",
                "grupo":"Gestión Clientes",
                "desc":"Relación entre lo que el banco presta y lo que capta en depósitos.",
                "ejemplo":"Colocaciones $80, Depósitos $100 → Ratio = 0.80.",
            },
        }

        col7, col8 = st.columns([0.75, 0.25])
        with col7:
            st.markdown('<span class="section-label">Seleccionar indicadores</span>', unsafe_allow_html=True)
        with col8:
            tog_r = st.toggle("Todos", value=st.session_state["all_r"], key="tog_r")
        if tog_r != st.session_state["all_r"]:
            st.session_state["all_r"] = tog_r; st.rerun()

        sel_ratios = st.multiselect("Indicadores Financieros:", list(M_CONF.keys()),
                                    default=list(M_CONF.keys()) if st.session_state["all_r"] else ["Liquidez Corriente"],
                                    label_visibility="collapsed")

        r_to_pdf = {}
        current_grupo = None
        for nom in sel_ratios:
            conf = M_CONF[nom]
            if conf.get("grupo") != current_grupo:
                current_grupo = conf.get("grupo")
                st.markdown(
                    f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};'
                    f'text-transform:uppercase;letter-spacing:1.2px;margin:32px 0 6px 0;">'
                    f'{current_grupo}</div>', unsafe_allow_html=True)

            desc    = conf.get("desc","")
            ejemplo = conf.get("ejemplo","")
            st.markdown(
                f'<div class="ratio-title-row">'
                f'<span class="ratio-title-text">{nom}</span>'
                f'<div class="info-popup-wrap"><span class="info-btn">?</span>'
                f'<div class="info-popup"><div class="info-popup-title">{nom}</div>'
                f'<div class="info-popup-desc">{desc}</div>'
                f'<div class="info-popup-example">Ejemplo: {ejemplo}</div>'
                f'</div></div></div>', unsafe_allow_html=True)

            r_data = []
            a_ref  = sel_años[0] if sel_años else 2024
            for b in sel_bancos:
                tmp     = df_b[(df_b["Banco"]==b)&(df_b["Anho"]==a_ref)]
                v1, v2  = tmp[tmp["Cuenta"]==conf["n1"]]["Valor"].sum(), tmp[tmp["Cuenta"]==conf["n2"]]["Valor"].sum()
                val     = v1/v2 if v2>0 else 0
                tmp_a   = df_b[(df_b["Banco"]==b)&(df_b["Anho"]==a_ref-1)]
                v1a,v2a = tmp_a[tmp_a["Cuenta"]==conf["n1"]]["Valor"].sum(), tmp_a[tmp_a["Cuenta"]==conf["n2"]]["Valor"].sum()
                val_ant = v1a/v2a if v2a>0 else 0
                var     = ((val-val_ant)/val_ant*100) if val_ant>0 else 0
                r_data.append({"b":b,"val":val,"var":var})

            target   = max([x["val"] for x in r_data]) if r_data else 0
            html_row = '<div class="ratio-row">'
            for res in r_data:
                is_l      = res["val"]==target and res["val"]>0
                icon      = "↑" if res["var"]>=0 else "↓"
                color_var = C["primary"] if res["var"]>=0 else "#576dff"
                bc        = BANK_COLOR.get(res["b"], C["primary"])
                html_row += (
                    f'<div class="ratio-card {"leader" if is_l else ""}">'
                    f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:2px;">'
                    f'<div style="width:3px;height:34px;border-radius:2px;background:{bc};flex-shrink:0;"></div>'
                    f'<div><div class="ratio-meta">{conf["f"]}</div>'
                    f'<div style="display:flex;align-items:center;gap:5px;">'
                    f'<span class="ratio-banco">{res["b"]}</span>'
                    f'{" <span class=\'badge-lider\'>LÍDER</span>" if is_l else ""}'
                    f'</div></div></div>'
                    f'<div class="ratio-divider"></div>'
                    f'<span class="ratio-num" style="color:{bc};">{res["val"]:.3f}</span>'
                    f'<div style="font-size:12px;font-weight:600;margin-top:4px;color:{color_var};">'
                    f'{icon} {abs(res["var"]):.1f}% vs {a_ref-1}</div>'
                    f'</div>'
                )
            st.markdown(html_row+'</div>', unsafe_allow_html=True)
            r_to_pdf[nom] = r_data

        # ── EXPORTAR ──────────────────────────────────────────────────────────
        st.markdown(
            f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;'
            f'letter-spacing:1.2px;margin:24px 0 10px 0;">Exportar</div>',
            unsafe_allow_html=True)
        col_r1, col_r2 = st.columns(2)

        with col_r1:
            if st.button("Generar PDF", use_container_width=True):
                ov = st.empty()
                ov.markdown(
                    f'<div class="minimal-load">'
                    f'<div class="spinner"></div>'
                    f'<div style="color:{C["primary"]};font-weight:600;font-size:13px;">Generando PDF...</div></div>',
                    unsafe_allow_html=True)
                try:
                    pdf_bytes = generar_pdf_profesional(df_filt, r_to_pdf, sel_bancos, sel_cuentas)
                    time.sleep(1); ov.empty()
                    st.download_button("⬇ Descargar PDF", data=pdf_bytes,
                                       file_name="Reporte_INSAIT.pdf",
                                       use_container_width=True)
                except Exception as e:
                    ov.empty(); st.error(f"Error: {e}")

        with col_r2:
            if st.button("Generar Excel", use_container_width=True):
                try:
                    import io, openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    wb   = openpyxl.Workbook()
                    # ── Hoja 1: Datos ─────────────────────────────────────
                    ws1  = wb.active; ws1.title = "Datos"
                    hdr_fill = PatternFill("solid", fgColor="4D6BFF")
                    hdr_font = Font(bold=True, color="FFFFFF", size=11)
                    thin = Border(
                        left=Side(style="thin",color="E5E7EB"),
                        right=Side(style="thin",color="E5E7EB"),
                        top=Side(style="thin",color="E5E7EB"),
                        bottom=Side(style="thin",color="E5E7EB"))
                    for ci, h in enumerate(["Banco","Año","Cuenta","Valor (MM$)"], 1):
                        c = ws1.cell(1, ci, h)
                        c.font = hdr_font; c.fill = hdr_fill
                        c.alignment = Alignment(horizontal="center"); c.border = thin
                    for _, row in df_filt.sort_values(["Banco","Anho","Cuenta"]).iterrows():
                        ws1.append([row["Banco"], row["Anho"], row["Cuenta"], round(row["Valor"],2)])
                        for ci in range(1, 5):
                            ws1.cell(ws1.max_row, ci).border = thin
                    for col in ws1.columns:
                        ws1.column_dimensions[get_column_letter(col[0].column)].width = \
                            max(len(str(c.value or "")) for c in col) + 4
                    # ── Hoja 2: Indicadores seleccionados ─────────────────
                    ws2 = wb.create_sheet("Indicadores")
                    for ci, h in enumerate(["Indicador","Grupo","Banco","Valor","Var % vs año ant."], 1):
                        c = ws2.cell(1, ci, h)
                        c.font = hdr_font; c.fill = hdr_fill
                        c.alignment = Alignment(horizontal="center"); c.border = thin
                    a_ref_xl = sel_años[0] if sel_años else 2024
                    for nom in sel_ratios:
                        rdata = r_to_pdf.get(nom, [])
                        grp   = M_CONF[nom].get("grupo","")
                        for res in rdata:
                            ws2.append([nom, grp, res["b"], round(res["val"],4), round(res["var"],2)])
                            for ci in range(1, 6):
                                ws2.cell(ws2.max_row, ci).border = thin
                    for col in ws2.columns:
                        ws2.column_dimensions[get_column_letter(col[0].column)].width = \
                            max(len(str(c.value or "")) for c in col) + 4
                    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                    st.download_button("⬇ Descargar Excel", data=buf.getvalue(),
                                       file_name="Reporte_INSAIT.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
                except Exception as e:
                    st.error(f"Error: {e}")

# ══════════════════════════════════════════════════════
#  EMPRESA
# ══════════════════════════════════════════════════════
with tab_empresa:
    st.info(f"Módulo de Empresas conectado a: {RUTA_BD}")

# ══════════════════════════════════════════════════════
#  MI EMPRESA
# ══════════════════════════════════════════════════════
with tab_mi_empresa:
    st.markdown(f'<span style="font-size:15px;color:{C["n600"]};">Módulo Mi Empresa — próximamente.</span>',
                unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
#  CONFIGURACION
# ══════════════════════════════════════════════════════
with tab_config:
    st.markdown(f'<span style="font-size:15px;color:{C["n600"]};">Configuracion — próximamente.</span>',
                unsafe_allow_html=True)