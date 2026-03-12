import streamlit as st
import pandas as pd
import requests
import plotly.graph_objects as go
import os, base64, time, tempfile, importlib.util
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from fpdf import FPDF

# ── CARGA VISTA MI EMPRESA ────────────────────────────────────────────────────
_me_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "views", "mi_empresa.py")
_me_spec = importlib.util.spec_from_file_location("mi_empresa", _me_path)
_me_mod  = importlib.util.module_from_spec(_me_spec)
_me_spec.loader.exec_module(_me_mod)
render_mi_empresa = _me_mod.render_mi_empresa

# ── PALETA ───────────────────────────────────────────────────────────────────
C = {
    "primary":       "#4d6bff",
    "primary_soft":  "rgba(77,107,255,0.08)",
    "n900":          "#0f172a",
    "n600":          "#475569",
    "n400":          "#94a3b8",
    "n100":          "#f1f5f9",
    "bg":            "#f8faff",
    "card":          "rgba(255,255,255,0.85)",
    "border":        "rgba(77,107,255,0.10)",
}

BRAND_COLORS = {
    "Banco de Chile":          "#0033A0",
    "Banco Internacional":     "#005EB8",
    "Banco Estado":            "#F5A800",
    "Scotiabank Chile":        "#ED1C24",
    "Bci":                     "#0057A8",
    "Banco Do Brasil":         "#F5A800",
    "JP Morgan Chase":         "#1A1A2E",
    "China Construction Bank": "#CC0000",
    "Santander Chile":         "#EC0000",
    "Itaú Corpbanca":          "#FF6200",
    "Banco Itaú":              "#FF6200",
    "BTG Pactual Chile":       "#003366",
    "Banco Security":          "#005BAA",
    "Banco Falabella":         "#00A651",
    "Banco BICE":              "#1A3A6C",
    "Banco Consorcio":         "#00457C",
    "Banco Ripley":            "#6B2D8B",
    "Banco ST":                "#2C7BE5",
}

st.set_page_config(page_title="INSAIT Pro", layout="wide", initial_sidebar_state="collapsed")

RUTA_INS_T = r"/Users/benjamin/Desktop/Tesis/INS T"
RUTA_BD    = r"C:\Users\Hp\Desktop\Tesis\BD"
API_KEY    = "ea9b378f0cbb6fc8b27040141e054e22752373f2"

# ── BANCO CENTRAL ─────────────────────────────────────────────────────────────
BCCH_USER = ""
BCCH_PASS = ""
BCCH_SERIES = {"tpm": "F022.BCH.INT.010.M", "uf": "F073.TCO.PRE.Z.D"}
TIPOS_CREDITO_MAP = {
    "Consumo": "consumo", "Hipotecario": "hipotecario",
    "Comercial": "comercial", "Leasing": "leasing",
}

# ── PDF ───────────────────────────────────────────────────────────────────────
class ReportePDF(FPDF):
    def header(self):
        logo_path = os.path.join(RUTA_INS_T, "1.png")
        if os.path.exists(logo_path): self.image(logo_path, x=10, y=8, w=30)
        self.set_font('Arial','B',10); self.set_text_color(150,150,150)
        self.cell(0,10,f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}',0,1,'R')
        self.ln(10); self.set_font('Arial','B',16); self.set_text_color(35,61,255)
        self.cell(0,10,'REPORTE FINANCIERO',0,1,'C'); self.ln(5)
    def footer(self):
        self.set_y(-15); self.set_font('Arial','I',8); self.set_text_color(128,128,128)
        self.cell(0,10,f'Página {self.page_no()}',0,0,'C')

def generar_pdf_profesional(df_filt, ratios_data, sel_bancos, sel_cuentas):
    pdf = ReportePDF(); pdf.set_auto_page_break(auto=True, margin=15); pdf.add_page()
    pdf.set_fill_color(35,61,255); pdf.set_text_color(255,255,255); pdf.set_font("Arial",'B',11)
    pdf.cell(0,10," 1. ANALISIS DE CUENTAS POR PERIODO (DATOS NOMINALES)",0,1,'L',True); pdf.ln(5)
    for banco in sel_bancos:
        df_banco = df_filt[df_filt["Banco"]==banco].sort_values(["Anho","Cuenta"])
        if df_banco.empty: continue
        pdf.set_text_color(35,61,255); pdf.set_font("Arial",'B',10)
        pdf.cell(0,8,f">> INSTITUCION: {banco}",0,1)
        pdf.set_fill_color(240,240,240); pdf.set_text_color(0,0,0); pdf.set_font("Arial",'B',9)
        pdf.cell(40,8,"Periodo",1,0,'C',True); pdf.cell(100,8,"Cuenta",1,0,'C',True)
        pdf.cell(50,8,"Valor (MM$)",1,1,'C',True); pdf.set_font("Arial",'',9)
        for _,row in df_banco.iterrows():
            pdf.cell(40,7,str(row['Anho']),1,0,'C'); pdf.cell(100,7,str(row['Cuenta']),1,0,'L')
            pdf.cell(50,7,f"{row['Valor']:,.0f}",1,1,'R')
        pdf.ln(5)
    pdf.add_page(); pdf.set_fill_color(35,61,255); pdf.set_text_color(255,255,255); pdf.set_font("Arial",'B',11)
    pdf.cell(0,10," 2. VISUALIZACION DE TENDENCIAS GRAFICAS",0,1,'L',True); pdf.ln(5)
    combos = [(b,ct) for b in sel_bancos for ct in sel_cuentas if not df_filt[(df_filt["Banco"]==b)&(df_filt["Cuenta"]==ct)].empty]
    with tempfile.TemporaryDirectory() as tmpdir:
        for idx,(b,ct) in enumerate(combos):
            d_p = df_filt[(df_filt["Banco"]==b)&(df_filt["Cuenta"]==ct)].sort_values("Anho")
            fig_p = go.Figure()
            fig_p.add_trace(go.Scatter(x=d_p["Anho"],y=d_p["Valor"],fill='tozeroy',fillcolor='rgba(35,61,255,0.1)',line=dict(color=C["primary"],width=4),mode='lines+markers'))
            fig_p.update_layout(title=f"{b}: {ct}",template="plotly_white",width=700,height=400)
            img_p = os.path.join(tmpdir,f"c_{idx}.png"); fig_p.write_image(img_p,scale=2)
            if pdf.get_y()>200: pdf.add_page()
            pdf.image(img_p,x=20,w=170); pdf.ln(2)
    return pdf.output(dest='S').encode('latin-1')

# ── DATOS ─────────────────────────────────────────────────────────────────────
def b64(p):
    try:
        if os.path.exists(p):
            with open(p,"rb") as f: return base64.b64encode(f.read()).decode()
    except: pass
    return None

def _fetch(c, a):
    u = f"https://api.cmfchile.cl/api-sbifv3/recursos_api/balances/{a}/12/instituciones/{c}?apikey={API_KEY}&formato=json"
    try:
        r = requests.get(u,timeout=12).json()
        return [{"Banco":c,"Anho":int(a),"Cuenta":i["DescripcionCuenta"].strip().upper(),"Valor":float(i["MonedaTotal"].replace(",","."))/1e6} for i in r.get("CodigosBalances",[])]
    except: return []

# ── FUNCIONES CRÉDITOS ────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def get_tasas_cmf(tipo: str) -> tuple:
    endpoints = [
        f"https://api.cmfchile.cl/api-sbifv3/recursos/v1/tasas/{tipo}",
        f"https://api.cmfchile.cl/api-sbifv3/recursos/v1/creditos/{tipo}/tasas",
        f"https://api.cmfchile.cl/api-sbifv3/recursos/v1/credito/{tipo}",
    ]
    last_error = ""
    for url in endpoints:
        try:
            r = requests.get(url, params={"apikey": API_KEY, "formato": "json"}, timeout=12)
            if r.status_code != 200:
                last_error = f"HTTP {r.status_code} · {url.split('v1/')[-1]}"; continue
            data = r.json()
            rows = []
            if isinstance(data, list): rows = data
            else:
                for key in ("Tasas","TasasCredito","TasasConsumo","TasasHipotecario","TasasComercial","TasasLeasing","Serie","Data"):
                    if key in data and isinstance(data[key], list): rows = data[key]; break
                if not rows:
                    for v in data.values():
                        if isinstance(v, list) and len(v) > 0: rows = v; break
            if rows: return rows, None
            last_error = "Endpoint retornó estructura vacía"
        except requests.exceptions.ConnectionError: last_error = "Sin conexión a la API CMF"
        except Exception as e: last_error = str(e)
    return [], last_error

@st.cache_data(ttl=3600, show_spinner=False)
def get_tasas_historicas_cmf(tipo: str, meses: int = 12) -> list:
    from datetime import date
    hoy = date.today(); registros = []
    for i in range(meses):
        mes = hoy.month - i; año = hoy.year
        while mes <= 0: mes += 12; año -= 1
        periodo = f"{año}{mes:02d}"
        try:
            url = f"https://api.cmfchile.cl/api-sbifv3/recursos/v1/tasas/{tipo}"
            r = requests.get(url, params={"apikey": API_KEY, "formato": "json", "periodo": periodo}, timeout=12)
            data = r.json(); rows = []
            if isinstance(data, list): rows = data
            else:
                for key in ("Tasas","TasasCredito","Serie"):
                    if key in data: rows = data[key]; break
            for row in rows: row["_periodo"] = periodo
            registros.extend(rows)
        except Exception: pass
    return registros

@st.cache_data(ttl=1800, show_spinner=False)
def get_tpm_bcch(start: str = "2023-01-01") -> list:
    if not BCCH_USER: return []
    try:
        params = {"user": BCCH_USER, "pass": BCCH_PASS, "function": "GetSeries", "timeseries": BCCH_SERIES["tpm"], "start": start, "end": datetime.today().strftime("%Y-%m-%d"), "format": "json"}
        r = requests.get("https://si3.bcentral.cl/SieteRestWS/SieteRestWS.ashx", params=params, timeout=12)
        obs = r.json()["Series"]["Obs"]
        return [{"fecha": o["indexDateString"], "tpm": float(o["value"])} for o in obs if o.get("value") not in (None, "N.D.", "")]
    except Exception: return []

def simular_cuota(monto: float, tasa_anual: float, plazo_meses: int) -> dict:
    if tasa_anual <= 0 or plazo_meses <= 0 or monto <= 0: return {}
    tm = (tasa_anual / 100) / 12
    cuota = monto * (tm * (1 + tm) ** plazo_meses) / ((1 + tm) ** plazo_meses - 1)
    total = cuota * plazo_meses
    return {"cuota": round(cuota), "total": round(total), "interes": round(total - monto), "costo_pct": round((total - monto) / monto * 100, 1)}

def _detectar_campos(rows: list) -> tuple:
    if not rows: return "Tasa", "Institucion"
    first = rows[0]
    campo_tasa  = next((k for k in ("Tasa","TasaAnual","tasa","TasaEfectiva") if k in first), list(first.keys())[0])
    campo_banco = next((k for k in ("Institucion","Banco","banco","Nombre","institucion") if k in first), list(first.keys())[1] if len(first) > 1 else list(first.keys())[0])
    return campo_tasa, campo_banco

@st.cache_data(show_spinner=False)
def cargar_bancos():
    m = {"001":"Banco de Chile","009":"Banco Internacional","012":"Banco Estado","014":"Scotiabank Chile","016":"Bci","027":"Banco Do Brasil","028":"JP Morgan Chase","031":"China Construction Bank","037":"Santander Chile","039":"Itaú Corpbanca","040":"Banco Itaú","041":"BTG Pactual Chile","049":"Banco Security","051":"Banco Falabella","053":"Banco BICE","055":"Banco Consorcio","057":"Banco Ripley","067":"Banco ST"}
    rows = []
    with ThreadPoolExecutor(max_workers=15) as ex:
        futs = [ex.submit(_fetch,c,a) for c in m.keys() for a in range(2021,2026)]
        for f in as_completed(futs): rows.extend(f.result())
    df = pd.DataFrame(rows)
    if not df.empty: df["Banco"] = df["Banco"].map(m)
    return df

# ── SESSION STATE ─────────────────────────────────────────────────────────────
if "cargado" not in st.session_state:
    logo_carga = b64(os.path.join(RUTA_INS_T, "1.png"))
    logo_html  = (f'<img src="data:image/png;base64,{logo_carga}" style="height:64px;margin-bottom:20px;filter:drop-shadow(0 4px 12px rgba(77,107,255,0.2));">' if logo_carga else f'<div style="font-size:22px;font-weight:800;color:{C["primary"]};margin-bottom:16px;">INSAIT Pro</div>')
    ph = st.empty()
    with ph.container():
        st.markdown(f'<div class="minimal-load">{logo_html}<div class="spinner"></div><div style="font-size:12px;font-weight:600;color:{C["n400"]};margin-top:12px;letter-spacing:0.5px;">Cargando datos...</div></div>', unsafe_allow_html=True)
    st.session_state["df_b"] = cargar_bancos()
    st.session_state["cargado"] = True
    time.sleep(0.8); ph.empty()

for k,v in [("all_b",False),("all_a",False),("all_c",False),("all_r",False)]:
    if k not in st.session_state: st.session_state[k] = v

df_b        = st.session_state["df_b"]
all_banks_s = sorted(df_b["Banco"].unique()) if not df_b.empty else []
BANK_COLOR  = {b: BRAND_COLORS.get(b, "#233dff") for b in all_banks_s}

# ── TABS NATIVOS ──────────────────────────────────────────────────────────────

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');

html, body, .stApp {{
    background-color: {C["bg"]} !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    color: {C["n900"]} !important;
}}
.stApp::before {{ content:''; position:fixed; top:-20%; left:-10%; width:60%; height:60%; background:radial-gradient(ellipse,rgba(77,107,255,0.08) 0%,transparent 70%); pointer-events:none; z-index:0; }}
.stApp::after  {{ content:''; position:fixed; bottom:-20%; right:-5%; width:50%; height:50%; background:radial-gradient(ellipse,rgba(77,107,255,0.05) 0%,transparent 70%); pointer-events:none; z-index:0; }}
*,label,p,div,span,h1,h2,h3,button,input,.stMarkdown {{ font-family:'Plus Jakarta Sans',sans-serif !important; }}
.main {{ display:flex !important; justify-content:center !important; }}
.main .block-container {{ width:100% !important; max-width:480px !important; min-width:0 !important; padding:1.2rem 1rem 80px 1rem !important; margin:0 auto !important; position:relative; z-index:1; }}
header[data-testid="stHeader"] {{ display:none !important; }}
footer {{ display:none !important; }}
#MainMenu {{ display:none !important; }}
[data-testid="stSidebar"] {{ display:none !important; }}
:root {{ --primary-color:{C["primary"]} !important; }}

[data-baseweb="checkbox"] [data-checked="true"],[data-baseweb="toggle"] [data-checked="true"],[data-testid="stToggle"] [aria-checked="true"],[role="switch"][aria-checked="true"] {{ background-color:{C["primary"]} !important; border-color:{C["primary"]} !important; }}
div[data-testid="stToggle"] label span {{ font-size:13px !important; color:{C["n600"]} !important; }}
span[data-baseweb="tag"] {{ background-color:{C["primary_soft"]} !important; border-color:rgba(77,107,255,0.2) !important; border-radius:8px !important; }}
span[data-baseweb="tag"] span {{ color:{C["primary"]} !important; font-weight:600 !important; }}
span[data-baseweb="tag"] svg {{ fill:{C["primary"]} !important; }}
[data-baseweb="menu"] li:hover,[data-baseweb="menu"] [aria-selected="true"] {{ background-color:{C["primary_soft"]} !important; color:{C["primary"]} !important; }}
div[data-baseweb="select"] > div {{ background:white !important; border:1.5px solid rgba(77,107,255,0.12) !important; border-radius:12px !important; }}

/* ── BOTONES GLOBALES ── */
.stButton > button,.stDownloadButton > button {{ background:linear-gradient(135deg,#4d6bff 0%,#3350e8 100%) !important; color:white !important; border:none !important; border-radius:14px !important; font-weight:700 !important; font-size:13px !important; padding:10px 20px !important; box-shadow:0 4px 14px rgba(77,107,255,0.22) !important; transition:all 0.2s !important; }}
.stButton > button p,.stDownloadButton > button p,.stButton > button span,.stDownloadButton > button span {{ color:white !important; }}
.stButton > button:hover,.stDownloadButton > button:hover {{ box-shadow:0 6px 20px rgba(77,107,255,0.35) !important; transform:translateY(-1px) !important; }}

/* ── RESET TOTAL: ningún stHorizontalBlock ni stColumn tiene fondo/borde/radius ── */
[data-testid="stHorizontalBlock"],
[data-testid="stColumn"] {{
    background: transparent !important;
    border: none !important;
    border-radius: 0 !important;
    padding: 0 !important;
    box-shadow: none !important;
    width: auto !important;
    margin: 0 !important;
}}

/* ── TABS NATIVOS — estilo pill compacto ── */
[data-baseweb="tab-list"] {{
    background: rgba(255,255,255,0.80) !important;
    border: 1.5px solid rgba(77,107,255,0.10) !important;
    border-radius: 99px !important;
    padding: 4px !important;
    gap: 2px !important;
    width: fit-content !important;
    margin: 0 auto 20px auto !important;
    box-shadow: 0 2px 12px rgba(77,107,255,0.08) !important;
    overflow-x: auto !important;
    scrollbar-width: none !important;
}}
[data-baseweb="tab-list"]::-webkit-scrollbar {{ display: none !important; }}
[data-baseweb="tab-highlight"],
[data-baseweb="tab-border"] {{
    display: none !important;
    opacity: 0 !important;
    height: 0 !important;
}}
[data-baseweb="tab"] {{
    background: transparent !important;
    border: none !important;
    border-radius: 99px !important;
    padding: 5px 12px !important;
    font-size: 11.5px !important;
    font-weight: 500 !important;
    color: {C["n600"]} !important;
    white-space: nowrap !important;
    min-height: 0 !important;
    height: auto !important;
    box-shadow: none !important;
    transition: all 0.15s !important;
    outline: none !important;
}}
[data-baseweb="tab"] p,
[data-baseweb="tab"] span {{
    font-size: 11.5px !important;
    font-weight: inherit !important;
    color: inherit !important;
    margin: 0 !important;
}}
[data-baseweb="tab"]:hover {{
    color: {C["primary"]} !important;
    background: rgba(77,107,255,0.06) !important;
}}
[data-baseweb="tab"][aria-selected="true"] {{
    background: {C["primary"]} !important;
    color: white !important;
    font-weight: 700 !important;
    box-shadow: 0 2px 8px rgba(77,107,255,0.28) !important;
}}
[data-baseweb="tab"][aria-selected="true"] p,
[data-baseweb="tab"][aria-selected="true"] span {{
    color: white !important;
    font-weight: 700 !important;
}}
[data-testid="stTabPanel"] {{
    padding: 0 !important;
}}

/* ── SLIDER AZUL ── */
[data-testid="stSlider"] [role="slider"] {{ background:{C["primary"]} !important; border-color:{C["primary"]} !important; }}
input[type="range"]::-webkit-slider-thumb {{ background:{C["primary"]} !important; }}
input[type="range"]::-webkit-slider-runnable-track {{ background:rgba(77,107,255,0.15) !important; }}

/* ── CARDS ── */
.modulo {{ background:white; border:1.5px solid rgba(77,107,255,0.08); border-radius:24px; padding:22px 22px 18px 22px; margin-bottom:14px; box-shadow:0 4px 24px rgba(77,107,255,0.07),0 1px 4px rgba(0,0,0,0.02); }}
.section-label {{ font-size:12px; font-weight:600; color:{C["n600"]}; letter-spacing:0.1px; margin-bottom:4px; display:block; }}
.ratio-row {{ display:flex; gap:12px; overflow-x:auto; padding:4px 2px 20px; flex-wrap:nowrap; -webkit-overflow-scrolling:touch; }}
.ratio-card {{ min-width:200px; background:white; border-radius:20px; padding:18px 16px; border:1.5px solid rgba(77,107,255,0.08); display:flex; flex-direction:column; gap:5px; box-shadow:0 4px 16px rgba(77,107,255,0.07); }}
.ratio-card.leader {{ border:1.5px solid rgba(77,107,255,0.35); background:rgba(77,107,255,0.04); box-shadow:0 4px 20px rgba(77,107,255,0.13); }}
.ratio-meta {{ font-size:9px; font-weight:700; color:{C["n400"]}; text-transform:uppercase; letter-spacing:1.2px; }}
.ratio-banco {{ font-size:12px; font-weight:600; color:{C["n900"]}; }}
.ratio-divider {{ height:1px; background:{C["n100"]}; margin:6px 0; }}
.ratio-num {{ font-size:28px; font-weight:800; letter-spacing:-1px; line-height:1.1; }}
.badge-lider {{ background:linear-gradient(135deg,{C["primary"]} 0%,#6b84ff 100%); color:white; font-size:8px; padding:2px 9px; border-radius:99px; font-weight:700; display:inline-flex; align-items:center; letter-spacing:0.5px; box-shadow:0 2px 8px rgba(77,107,255,0.3); }}
.ratio-title-row {{ display:flex; align-items:center; gap:8px; margin:8px 0 10px 0; }}
.ratio-title-text {{ font-weight:700; color:{C["n900"]}; font-size:14px; letter-spacing:-0.2px; }}
.info-popup-wrap {{ position:relative; display:inline-flex; align-items:center; }}
.info-btn {{ display:inline-flex; align-items:center; justify-content:center; width:17px; height:17px; border-radius:50%; background:{C["n100"]}; color:{C["n400"]}; font-size:10px; font-weight:700; cursor:default; user-select:none; flex-shrink:0; transition:all 0.15s; border:1px solid rgba(77,107,255,0.10); }}
.info-btn:hover {{ background:{C["primary_soft"]}; color:{C["primary"]}; }}
.info-popup {{ display:none; position:absolute; left:24px; top:50%; transform:translateY(-50%); z-index:9999; background:white; border:1.5px solid rgba(77,107,255,0.14); border-radius:18px; padding:16px 18px; width:270px; box-shadow:0 12px 40px rgba(77,107,255,0.12),0 2px 8px rgba(0,0,0,0.05); pointer-events:none; }}
.info-popup-wrap:hover .info-popup {{ display:block; }}
.info-popup-title {{ font-size:12px; font-weight:700; color:{C["primary"]}; margin-bottom:6px; }}
.info-popup-desc {{ font-size:11px; color:{C["n600"]}; line-height:1.6; margin-bottom:10px; }}
.info-popup-example {{ background:{C["primary_soft"]}; border:1px solid rgba(77,107,255,0.15); border-radius:10px; padding:8px 10px; font-size:10px; color:{C["primary"]}; font-weight:600; line-height:1.5; }}

/* ── ELIMINAR CONTENEDOR REDONDEADO DE FONDO (stVerticalBlockBorderWrapper) ── */
[data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stVerticalBlockBorderWrapper"] > div {{
    background: transparent !important;
    border: none !important;
    border-radius: 0 !important;
    box-shadow: none !important;
    padding: 0 !important;
}}

/* ── RADIO: botones planos uno al lado del otro, solo seleccionado se colorea ── */
div[data-testid="stRadio"] > div[role="radiogroup"] {{
    background: transparent !important;
    border: none !important;
    border-radius: 0 !important;
    padding: 0 !important;
    display: flex !important;
    flex-direction: row !important;
    gap: 8px !important;
    box-shadow: none !important;
    width: fit-content !important;
}}
div[data-testid="stRadio"] input[type="radio"] {{ position:absolute !important; opacity:0 !important; width:0 !important; height:0 !important; pointer-events:none !important; }}
div[data-testid="stRadio"] label {{
    display:flex !important; align-items:center !important;
    border-radius:10px !important; padding:6px 16px !important;
    font-size:12px !important; font-weight:500 !important;
    color:{C["n600"]} !important; cursor:pointer !important;
    transition:all 0.15s !important;
    background:transparent !important;
    border:1.5px solid transparent !important;
    white-space:nowrap !important; gap:0 !important;
}}
div[data-testid="stRadio"] label > div:first-child {{ display:none !important; }}
div[data-testid="stRadio"] label:hover {{
    color:{C["primary"]} !important;
    border-color:rgba(77,107,255,0.20) !important;
}}
div[data-testid="stRadio"] label:has(input:checked) {{
    background:{C["primary"]} !important; color:white !important;
    font-weight:700 !important; border-color:{C["primary"]} !important;
    box-shadow:0 2px 10px rgba(77,107,255,0.25) !important;
}}
div[data-testid="stRadio"] label:has(input:checked) p,
div[data-testid="stRadio"] label:has(input:checked) span {{ color:white !important; }}
div[data-testid="stRadio"] label p {{ font-size:12px !important; font-weight:inherit !important; color:inherit !important; margin:0 !important; }}

/* ── LOADER ── */
.minimal-load {{ position:fixed; inset:0; background:rgba(248,250,255,0.94); backdrop-filter:blur(12px); z-index:99999; display:flex; flex-direction:column; align-items:center; justify-content:center; }}
.spinner {{ width:28px; height:28px; border:2px solid rgba(77,107,255,0.15); border-top:2px solid {C["primary"]}; border-radius:50%; animation:spin 0.85s linear infinite; margin-bottom:14px; }}
@keyframes spin {{ to {{ transform:rotate(360deg); }} }}

/* ── ELIMINAR DOTS DECORATIVOS RESIDUALES ── */
div[style*="border-radius:50%"][style*="width:7px"],
div[style*="border-radius: 50%"][style*="width: 7px"] {{
    display: none !important;
}}



</style>
""", unsafe_allow_html=True)

# ── RENDER TABS ───────────────────────────────────────────────────────────────
tab_inicio, tab_banco, tab_empresa, tab_mi_empresa, tab_config = st.tabs(
    ["Inicio", "Banco", "Empresa", "Mi Empresa", "Configuración"]
)

# ══════════════════════════════════════════════════════
#  INICIO
# ══════════════════════════════════════════════════════
with tab_inicio:
    import datetime as dt

    @st.cache_data(ttl=3600)
    def get_indicadores():
        try:
            r = requests.get("https://mindicador.cl/api", timeout=5); d = r.json()
            return {"dolar":d.get("dolar",{}).get("valor","N/A"),"uf":d.get("uf",{}).get("valor","N/A"),"euro":d.get("euro",{}).get("valor","N/A"),"utm":d.get("utm",{}).get("valor","N/A"),"ipc":d.get("ipc",{}).get("valor","N/A"),"tpm":d.get("tpm",{}).get("valor","N/A")}
        except: return {"dolar":"—","uf":"—","euro":"—","utm":"—","ipc":"—","tpm":"—"}

    def fmtv(v, pref="$", suf=""):
        if v in ("N/A","—"): return "—"
        try: return f"{pref}{float(v):,.2f}{suf}"
        except: return str(v)

    ind  = get_indicadores()
    hoy  = dt.date.today().strftime("%d de %B, %Y")
    hora = dt.datetime.now().strftime("%H:%M")

    st.markdown(f"""
    <div style="padding:18px 20px;background:white;border-radius:20px;
                border:1.5px solid rgba(77,107,255,0.08);
                box-shadow:0 2px 10px rgba(77,107,255,0.05);margin-bottom:20px;">
      <div style="font-size:10px;font-weight:700;color:{C["n400"]};text-transform:uppercase;letter-spacing:1.4px;margin-bottom:4px;">BIENVENIDO</div>
      <div style="font-size:21px;font-weight:800;color:{C["n900"]};letter-spacing:-0.5px;line-height:1.2;">Hola, INSAIT 👋</div>
      <div style="font-size:11px;color:{C["n400"]};margin-top:3px;">{hoy} · {hora}</div>
    </div>""", unsafe_allow_html=True)

    st.markdown(f'<div style="font-size:13px;font-weight:700;color:{C["n900"]};margin-bottom:12px;letter-spacing:-0.2px;">Indicadores del día</div>', unsafe_allow_html=True)

    ind_items = [
        ("Dólar",  fmtv(ind["dolar"]),               C["primary"], "USD"),
        ("UF",     fmtv(ind["uf"]),                  "#0057A8",    "Unidad de Fomento"),
        ("Euro",   fmtv(ind["euro"]),                "#005EB8",    "EUR"),
        ("TPM",    fmtv(ind["tpm"],pref="",suf="%"), "#06b6a0",    "Banco Central"),
        ("UTM",    fmtv(ind["utm"]),                 "#6B2D8B",    "Unid. Trib. Mensual"),
        ("IVA",    "19%",                            "#475569",    "Impuesto al valor"),
    ]
    for row_start in range(0, len(ind_items), 2):
        ca, cb = st.columns(2, gap="small")
        for col, item in zip([ca, cb], ind_items[row_start:row_start+2]):
            label, valor, color, sublabel = item
            with col:
                st.markdown(f"""
                <div style="background:white;border:1.5px solid rgba(77,107,255,0.07);border-radius:18px;padding:16px 16px 14px;margin-bottom:10px;box-shadow:0 2px 8px rgba(77,107,255,0.04);">
                  <div style="font-size:10px;font-weight:700;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;margin-bottom:6px;">{label}</div>
                  <div style="font-size:19px;font-weight:800;color:{C["n900"]};letter-spacing:-0.5px;line-height:1;">{valor}</div>
                  <div style="font-size:10px;color:{C["n400"]};margin-top:4px;">{sublabel}</div>
                </div>""", unsafe_allow_html=True)

    st.markdown(f'<div style="font-size:13px;font-weight:700;color:{C["n900"]};margin:4px 0 12px;letter-spacing:-0.2px;">Noticias recientes</div>', unsafe_allow_html=True)
    noticias = [
        (C["primary"], "MERCADO",  "Proyecciones económicas para el segundo semestre", "Hace 2 horas"),
        (C["n600"],    "EMPRESAS", "Nuevas normativas para facturación electrónica",   "Hace 5 horas"),
        ("#06b6a0",    "BANCO",    "Tasas de interés se mantienen estables en Chile",  "Hace 8 horas"),
    ]
    for color, cat, titulo, tiempo in noticias:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:14px;background:white;border:1.5px solid rgba(77,107,255,0.07);border-radius:16px;padding:14px 16px;margin-bottom:9px;box-shadow:0 2px 8px rgba(77,107,255,0.04);">
          <div style="width:4px;height:40px;border-radius:2px;background:{color};flex-shrink:0;"></div>
          <div style="flex:1;min-width:0;">
            <div style="font-size:9px;font-weight:700;color:{color};text-transform:uppercase;letter-spacing:1.2px;margin-bottom:3px;">{cat}</div>
            <div style="font-size:12px;font-weight:600;color:{C["n900"]};line-height:1.4;margin-bottom:3px;">{titulo}</div>
            <div style="font-size:10px;color:{C["n400"]};">{tiempo}</div>
          </div>
        </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
#  BANCO
# ══════════════════════════════════════════════════════
with tab_banco:

    sub_tab = st.radio("", ["Análisis Financiero", "Simulación de Créditos"], horizontal=True, label_visibility="collapsed", key="banco_subtab")
    st.markdown("<hr style='border:none;border-top:1.5px solid rgba(77,107,255,0.08);margin:12px 0 16px 0;'>", unsafe_allow_html=True)

    if sub_tab == "Análisis Financiero":
        b_l = sorted(df_b["Banco"].unique())

        with st.container():
            st.markdown('<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;"><span style="font-size:12px;font-weight:600;color:#64748b;">Bancos</span></div>', unsafe_allow_html=True)
            tog_b = st.toggle("Todos", value=st.session_state["all_b"], key="tog_b")
            if tog_b != st.session_state["all_b"]: st.session_state["all_b"] = tog_b; st.rerun()
            sel_bancos = st.multiselect("Bancos", b_l, default=b_l if st.session_state["all_b"] else [], label_visibility="collapsed")

            a_l = sorted(df_b["Anho"].unique(), reverse=True)
            st.markdown('<div style="margin-top:8px;margin-bottom:4px;"><span style="font-size:12px;font-weight:600;color:#64748b;">Periodos</span></div>', unsafe_allow_html=True)
            tog_a = st.toggle("Todos", value=st.session_state["all_a"], key="tog_a")
            if tog_a != st.session_state["all_a"]: st.session_state["all_a"] = tog_a; st.rerun()
            sel_años = st.multiselect("Periodos", a_l, default=a_l if st.session_state["all_a"] else [], label_visibility="collapsed")

            c_std = ["TOTAL ACTIVOS","TOTAL PASIVOS","PATRIMONIO","UTILIDAD (PÉRDIDA) DEL EJERCICIO","COLOCACIONES NETAS","DEPOSITOS Y OTRAS CAPTACIONES","GASTOS OPERACIONALES","INGRESOS OPERACIONALES","COLOCACIONES VENCIDAS","PROVISIONES","INGRESOS FINANCIEROS NETOS"]
            st.markdown('<div style="margin-top:8px;margin-bottom:4px;"><span style="font-size:12px;font-weight:600;color:#64748b;">Cuentas</span></div>', unsafe_allow_html=True)
            tog_c = st.toggle("Todos", value=st.session_state["all_c"], key="tog_c")
            if tog_c != st.session_state["all_c"]: st.session_state["all_c"] = tog_c; st.rerun()
            sel_cuentas = st.multiselect("Cuentas", c_std, default=c_std if st.session_state["all_c"] else [], label_visibility="collapsed")

        df_filt = df_b[(df_b["Banco"].isin(sel_bancos))&(df_b["Cuenta"].isin(sel_cuentas))&(df_b["Anho"].isin(sel_años))]

        if sel_bancos or sel_años or sel_cuentas:
            total_esperado   = len(sel_bancos)*len(sel_años)*len(sel_cuentas) if (sel_bancos and sel_años and sel_cuentas) else 0
            total_real       = len(df_filt[["Banco","Anho","Cuenta"]].drop_duplicates()) if not df_filt.empty else 0
            pct_cobertura    = int(total_real/total_esperado*100) if total_esperado>0 else 0
            bancos_sin_datos = [b for b in sel_bancos if df_filt[df_filt["Banco"]==b].empty] if sel_bancos else []
            total_registros  = len(df_filt) if not df_filt.empty else 0
            if pct_cobertura>=90:   calidad_color,calidad_icon,calidad_txt="#06d6a0","✓",f"Completos ({pct_cobertura}%)"
            elif pct_cobertura>=60: calidad_color,calidad_icon,calidad_txt="#F5A800","⚠",f"Parciales ({pct_cobertura}%)"
            else:                   calidad_color,calidad_icon,calidad_txt="#e63946","✗",f"Insuficientes ({pct_cobertura}%)"
            aviso_sin = f" · Sin datos: {', '.join(bancos_sin_datos)}" if bancos_sin_datos else ""
            st.markdown(f'<div class="modulo" style="padding:14px 20px;"><div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;"><div style="display:flex;gap:28px;align-items:center;"><div style="text-align:center;"><div style="font-size:22px;font-weight:700;color:{C["n900"]};line-height:1;">{len(sel_bancos)}</div><div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Bancos</div></div><div style="width:1px;height:28px;background:{C["n100"]};"></div><div style="text-align:center;"><div style="font-size:22px;font-weight:700;color:{C["n900"]};line-height:1;">{len(sel_años)}</div><div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Periodos</div></div><div style="width:1px;height:28px;background:{C["n100"]};"></div><div style="text-align:center;"><div style="font-size:22px;font-weight:700;color:{C["n900"]};line-height:1;">{len(sel_cuentas)}</div><div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Cuentas</div></div><div style="width:1px;height:28px;background:{C["n100"]};"></div><div style="text-align:center;"><div style="font-size:22px;font-weight:700;color:{C["primary"]};line-height:1;">{total_registros:,}</div><div style="font-size:10px;color:{C["n400"]};text-transform:uppercase;letter-spacing:0.8px;font-weight:600;margin-top:2px;">Registros</div></div></div><div style="display:inline-flex;align-items:center;gap:5px;background:{calidad_color}18;border:1px solid {calidad_color}44;border-radius:8px;padding:5px 10px;"><span style="color:{calidad_color};font-weight:700;font-size:12px;">{calidad_icon}</span><span style="color:{calidad_color};font-size:11px;font-weight:600;">{calidad_txt}{aviso_sin}</span></div></div></div>', unsafe_allow_html=True)

        st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:20px 0 10px 0;">Gráfico</div>', unsafe_allow_html=True)
        vista_lineas = st.toggle("Líneas", value=False, key="vista_toggle")
        fig = go.Figure(); v_t = "Líneas" if vista_lineas else "Barras"
        if not df_filt.empty and sel_cuentas:
            for b in sel_bancos:
                for cuenta in sel_cuentas:
                    d = df_filt[(df_filt["Banco"]==b)&(df_filt["Cuenta"]==cuenta)].sort_values("Anho")
                    if not d.empty:
                        bc  = BANK_COLOR.get(b, C["primary"])
                        lbl = f"{b} — {cuenta.title()}" if len(sel_cuentas)>1 else b
                        if v_t=="Barras": fig.add_trace(go.Bar(x=d["Anho"],y=d["Valor"],name=lbl,marker_color=bc,marker_line_width=0))
                        else: fig.add_trace(go.Scatter(x=d["Anho"],y=d["Valor"],name=lbl,mode='lines+markers',line=dict(color=bc,width=2.5),marker=dict(color=bc,size=7)))
            tit = ", ".join([c.title() for c in sel_cuentas])
        else:
            fig.add_trace(go.Scatter(x=[2021,2022,2023,2024,2025],y=[30,45,38,58,52],fill="tozeroy",fillcolor="rgba(35,61,255,0.07)",line=dict(color="rgba(35,61,255,0.22)",width=2,dash="dot"),mode="lines",hoverinfo="skip",showlegend=False))
            fig.add_annotation(text="<b>Seleccione datos para visualizar el análisis</b>",xref="paper",yref="paper",x=0.5,y=0.62,showarrow=False,font=dict(size=15,color="rgba(35,61,255,0.50)",family="DM Sans"))
            fig.add_annotation(text="Elija un banco, periodo y cuenta para comenzar",xref="paper",yref="paper",x=0.5,y=0.50,showarrow=False,font=dict(size=12,color="rgba(75,85,99,0.45)",family="DM Sans"))
            tit = "Vista Previa"
        fig.update_layout(title=dict(text=tit,font=dict(size=13,color=C["n600"],family="DM Sans")),template="plotly_white",height=320,margin=dict(l=0,r=0,t=40,b=10),legend=dict(orientation="h",y=-0.35,font=dict(family="DM Sans",size=11)),barmode="group")
        st.plotly_chart(fig, use_container_width=True)

        if sel_bancos and sel_años and not df_filt.empty:
            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:20px 0 10px 0;">Ranking</div>', unsafe_allow_html=True)
            todas_cuentas = sorted(df_filt["Cuenta"].unique().tolist())
            cuenta_rank = st.selectbox("Cuenta para ranking", todas_cuentas, key="cuenta_rank_sel", label_visibility="collapsed")
            a_rank  = sel_años[0]
            df_rank = df_filt[(df_filt["Cuenta"]==cuenta_rank)&(df_filt["Anho"]==a_rank)].sort_values("Valor",ascending=False).reset_index(drop=True)
            if not df_rank.empty:
                st.markdown(f'<div style="font-size:12px;color:{C["n400"]};margin-bottom:12px;">{cuenta_rank.title()} · {a_rank}</div>', unsafe_allow_html=True)
                medals = ["🥇","🥈","🥉"]; max_val = df_rank["Valor"].max() or 1
                html_rank = '<div style="display:flex;flex-direction:column;gap:7px;margin-bottom:4px;">'
                for i, row in df_rank.iterrows():
                    pct = row["Valor"]/max_val*100 if max_val>0 else 0
                    med = medals[i] if i<3 else f'<span style="font-size:11px;color:{C["n400"]};min-width:22px;display:inline-block;text-align:center;">#{i+1}</span>'
                    bc  = BANK_COLOR.get(row["Banco"], C["primary"])
                    html_rank += f'<div style="display:flex;align-items:center;gap:10px;"><span style="font-size:15px;min-width:22px;text-align:center;">{med}</span><div style="flex:1;"><div style="display:flex;justify-content:space-between;margin-bottom:3px;"><span style="font-size:12px;font-weight:600;color:{C["n900"]};">{row["Banco"]}</span><span style="font-size:12px;font-weight:700;color:{bc};">{row["Valor"]:,.0f} MM$</span></div><div style="background:{C["n100"]};border-radius:99px;height:5px;"><div style="background:{bc};width:{pct:.1f}%;height:5px;border-radius:99px;"></div></div></div></div>'
                st.markdown(html_rank+'</div>', unsafe_allow_html=True)

        if sel_bancos:
            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:20px 0 10px 0;">Indicadores Financieros</div>', unsafe_allow_html=True)
            M_CONF = {
                "Liquidez Corriente":     {"f":"ACTIVO / PASIVO","n1":"TOTAL ACTIVOS","n2":"TOTAL PASIVOS","grupo":"Financieros","desc":"Mide cuántos activos tiene el banco por cada peso de deuda. Valores mayores a 1 indican solidez.","ejemplo":"Si el banco tiene $200 en activos y $100 en pasivos → Liquidez = 2.0."},
                "Solvencia":              {"f":"PATRIMONIO / PASIVO","n1":"PATRIMONIO","n2":"TOTAL PASIVOS","grupo":"Financieros","desc":"Indica cuánto patrimonio propio respalda las deudas del banco.","ejemplo":"Patrimonio $50, Pasivos $200 → Solvencia = 0.25."},
                "ROA":                    {"f":"UTILIDAD / ACTIVOS","n1":"UTILIDAD (PÉRDIDA) DEL EJERCICIO","n2":"TOTAL ACTIVOS","grupo":"Financieros","desc":"Mide la rentabilidad del banco en relación a sus activos totales.","ejemplo":"Utilidad $10, Activos $500 → ROA = 0.02 (2%)."},
                "ROE":                    {"f":"UTILIDAD / PATRIMONIO","n1":"UTILIDAD (PÉRDIDA) DEL EJERCICIO","n2":"PATRIMONIO","grupo":"Financieros","desc":"Rentabilidad sobre el capital de los accionistas.","ejemplo":"Utilidad $10, Patrimonio $80 → ROE = 0.125 (12.5%)."},
                "NIM":                    {"f":"ING. FINANCIEROS NETOS / ACTIVOS","n1":"INGRESOS FINANCIEROS NETOS","n2":"TOTAL ACTIVOS","grupo":"Financieros","desc":"Margen Financiero Neto sobre activos.","ejemplo":"Ingresos netos $15, Activos $500 → NIM = 0.03 (3%)."},
                "Ratio de Eficiencia":    {"f":"GASTOS / INGRESOS","n1":"GASTOS OPERACIONALES","n2":"INGRESOS OPERACIONALES","grupo":"Financieros","desc":"Cuánto gasta el banco por cada peso que ingresa. Menor = más eficiente.","ejemplo":"Gastos $60, Ingresos $100 → Eficiencia = 0.60."},
                "Índice de Morosidad":    {"f":"COL. VENCIDAS / COL. TOTAL","n1":"COLOCACIONES VENCIDAS","n2":"COLOCACIONES NETAS","grupo":"Calidad de Activos y Riesgos","desc":"Porcentaje de créditos que no se están pagando a tiempo.","ejemplo":"Vencidos $5, Total $100 → Morosidad = 0.05 (5% impagos)."},
                "Índice de Riesgo":       {"f":"PROVISIONES / COLOCACIONES","n1":"PROVISIONES","n2":"COLOCACIONES NETAS","grupo":"Calidad de Activos y Riesgos","desc":"Cuánto ha reservado el banco para cubrir posibles pérdidas.","ejemplo":"Provisiones $8, Colocaciones $100 → Riesgo = 0.08."},
                "Ratio de Capital":       {"f":"PASIVO / PATRIMONIO","n1":"TOTAL PASIVOS","n2":"PATRIMONIO","grupo":"Calidad de Activos y Riesgos","desc":"Nivel de endeudamiento relativo al capital propio.","ejemplo":"Pasivos $300, Patrimonio $100 → Ratio = 3.0."},
                "Participación de Mercado":{"f":"COL. BANCO / DEPÓSITOS","n1":"COLOCACIONES NETAS","n2":"DEPOSITOS Y OTRAS CAPTACIONES","grupo":"Gestión Clientes","desc":"Relación entre lo que el banco presta y lo que capta en depósitos.","ejemplo":"Colocaciones $80, Depósitos $100 → Ratio = 0.80."},
            }
            col7,col8 = st.columns([0.75,0.25])
            with col7: st.markdown('<span class="section-label">Seleccionar indicadores</span>', unsafe_allow_html=True)
            with col8: tog_r = st.toggle("Todos", value=st.session_state["all_r"], key="tog_r")
            if tog_r != st.session_state["all_r"]: st.session_state["all_r"] = tog_r; st.rerun()
            sel_ratios = st.multiselect("Indicadores Financieros:", list(M_CONF.keys()), default=list(M_CONF.keys()) if st.session_state["all_r"] else ["Liquidez Corriente"], label_visibility="collapsed")
            r_to_pdf = {}; current_grupo = None
            for nom in sel_ratios:
                conf = M_CONF[nom]
                if conf.get("grupo") != current_grupo:
                    current_grupo = conf.get("grupo")
                    st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:32px 0 6px 0;">{current_grupo}</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="ratio-title-row"><span class="ratio-title-text">{nom}</span><div class="info-popup-wrap"><span class="info-btn">?</span><div class="info-popup"><div class="info-popup-title">{nom}</div><div class="info-popup-desc">{conf.get("desc","")}</div><div class="info-popup-example">Ejemplo: {conf.get("ejemplo","")}</div></div></div></div>', unsafe_allow_html=True)
                r_data = []; a_ref = sel_años[0] if sel_años else 2024
                for b in sel_bancos:
                    tmp=df_b[(df_b["Banco"]==b)&(df_b["Anho"]==a_ref)]; v1,v2=tmp[tmp["Cuenta"]==conf["n1"]]["Valor"].sum(),tmp[tmp["Cuenta"]==conf["n2"]]["Valor"].sum(); val=v1/v2 if v2>0 else 0
                    tmp_a=df_b[(df_b["Banco"]==b)&(df_b["Anho"]==a_ref-1)]; v1a,v2a=tmp_a[tmp_a["Cuenta"]==conf["n1"]]["Valor"].sum(),tmp_a[tmp_a["Cuenta"]==conf["n2"]]["Valor"].sum(); val_ant=v1a/v2a if v2a>0 else 0
                    var=((val-val_ant)/val_ant*100) if val_ant>0 else 0; r_data.append({"b":b,"val":val,"var":var})
                target=max([x["val"] for x in r_data]) if r_data else 0; html_row='<div class="ratio-row">'
                for res in r_data:
                    is_l=res["val"]==target and res["val"]>0; icon="↑" if res["var"]>=0 else "↓"; color_var=C["primary"] if res["var"]>=0 else "#576dff"; bc=BANK_COLOR.get(res["b"],C["primary"])
                    html_row += f'<div class="ratio-card {"leader" if is_l else ""}"><div style="display:flex;align-items:center;gap:8px;margin-bottom:2px;"><div style="width:3px;height:34px;border-radius:2px;background:{bc};flex-shrink:0;"></div><div><div class="ratio-meta">{conf["f"]}</div><div style="display:flex;align-items:center;gap:5px;"><span class="ratio-banco">{res["b"]}</span>{" <span class=\'badge-lider\'>LÍDER</span>" if is_l else ""}</div></div></div><div class="ratio-divider"></div><span class="ratio-num" style="color:{bc};">{res["val"]:.3f}</span><div style="font-size:12px;font-weight:600;margin-top:4px;color:{color_var};">{icon} {abs(res["var"]):.1f}% vs {a_ref-1}</div></div>'
                st.markdown(html_row+'</div>', unsafe_allow_html=True); r_to_pdf[nom]=r_data

            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:24px 0 10px 0;">Exportar</div>', unsafe_allow_html=True)
            col_r1,col_r2 = st.columns(2)
            with col_r1:
                if st.button("Generar PDF", use_container_width=True):
                    ov = st.empty()
                    ov.markdown(f'<div class="minimal-load"><div class="spinner"></div><div style="color:{C["primary"]};font-weight:600;font-size:13px;">Generando PDF...</div></div>', unsafe_allow_html=True)
                    try:
                        pdf_bytes = generar_pdf_profesional(df_filt, r_to_pdf, sel_bancos, sel_cuentas)
                        time.sleep(1); ov.empty()
                        st.download_button("⬇ Descargar PDF", data=pdf_bytes, file_name="Reporte_INSAIT.pdf", use_container_width=True)
                    except Exception as e: ov.empty(); st.error(f"Error: {e}")
            with col_r2:
                if st.button("Generar Excel", use_container_width=True):
                    try:
                        import io, openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        from openpyxl.utils import get_column_letter
                        wb=openpyxl.Workbook(); ws1=wb.active; ws1.title="Datos"
                        hdr_fill=PatternFill("solid",fgColor="4D6BFF"); hdr_font=Font(bold=True,color="FFFFFF",size=11)
                        thin=Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))
                        for ci,h in enumerate(["Banco","Año","Cuenta","Valor (MM$)"],1): c=ws1.cell(1,ci,h); c.font=hdr_font; c.fill=hdr_fill; c.alignment=Alignment(horizontal="center"); c.border=thin
                        for _,row in df_filt.sort_values(["Banco","Anho","Cuenta"]).iterrows():
                            ws1.append([row["Banco"],row["Anho"],row["Cuenta"],round(row["Valor"],2)])
                            for ci in range(1,5): ws1.cell(ws1.max_row,ci).border=thin
                        for col in ws1.columns: ws1.column_dimensions[get_column_letter(col[0].column)].width=max(len(str(c.value or "")) for c in col)+4
                        ws2=wb.create_sheet("Indicadores")
                        for ci,h in enumerate(["Indicador","Grupo","Banco","Valor","Var % vs año ant."],1): c=ws2.cell(1,ci,h); c.font=hdr_font; c.fill=hdr_fill; c.alignment=Alignment(horizontal="center"); c.border=thin
                        for nom in sel_ratios:
                            rdata=r_to_pdf.get(nom,[]); grp=M_CONF[nom].get("grupo","")
                            for res in rdata: ws2.append([nom,grp,res["b"],round(res["val"],4),round(res["var"],2)]); [setattr(ws2.cell(ws2.max_row,ci),'border',thin) for ci in range(1,6)]
                        for col in ws2.columns: ws2.column_dimensions[get_column_letter(col[0].column)].width=max(len(str(c.value or "")) for c in col)+4
                        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
                        st.download_button("⬇ Descargar Excel",data=buf.getvalue(),file_name="Reporte_INSAIT.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
                    except Exception as e: st.error(f"Error: {e}")

    if sub_tab == "Simulación de Créditos":
        tipo_label = st.selectbox("Tipo de crédito", list(TIPOS_CREDITO_MAP.keys()), key="tipo_cred_sel", label_visibility="collapsed")
        tipo_api   = TIPOS_CREDITO_MAP[tipo_label]
        with st.spinner("Consultando API CMF..."): tasas_actuales, cmf_error = get_tasas_cmf(tipo_api)

        tpm_data = get_tpm_bcch()
        if tpm_data:
            ultimo_tpm = tpm_data[-1]
            st.markdown(f"""<div style="background:linear-gradient(135deg,#eef1ff,#f8faff);border:1.5px solid rgba(77,107,255,0.15);border-radius:18px;padding:14px 20px;display:flex;gap:28px;align-items:center;margin-bottom:16px;flex-wrap:wrap;"><div><div style="font-size:24px;font-weight:800;color:{C["primary"]};letter-spacing:-1px;">{ultimo_tpm['tpm']}%</div><div style="font-size:9px;font-weight:700;color:{C["n400"]};text-transform:uppercase;letter-spacing:1px;">TPM Banco Central</div></div><div style="font-size:11px;color:{C["n600"]};">Período: {ultimo_tpm['fecha']}</div></div>""", unsafe_allow_html=True)
        elif not BCCH_USER:
            st.markdown(f"""<div style="background:#f8faff;border:1.5px solid rgba(77,107,255,0.10);border-radius:14px;padding:10px 16px;margin-bottom:12px;font-size:11px;color:{C["n600"]};">💡 Agrega <b>BCCH_USER</b> y <b>BCCH_PASS</b> para ver la TPM. Registro gratuito en <a href="https://si3.bcentral.cl/Secure/Users/Login.aspx" target="_blank" style="color:{C["primary"]};">si3.bcentral.cl</a></div>""", unsafe_allow_html=True)

        if not tasas_actuales:
            detalle = cmf_error or "Endpoint no disponible"
            st.markdown(f"""
            <div style="background:#fff7ed;border:1.5px solid #fed7aa;border-radius:16px;padding:16px 18px;margin-bottom:8px;">
              <div style="font-size:12px;font-weight:700;color:#9a3412;margin-bottom:6px;">⚠️ Sin datos de tasas para {tipo_label}</div>
              <div style="font-size:11px;color:#c2410c;margin-bottom:10px;">{detalle}</div>
              <div style="font-size:11px;color:#92400e;line-height:1.6;">
                Verifica el endpoint directamente:<br>
                <a href="https://api.cmfchile.cl/api-sbifv3/recursos/v1/tasas/{tipo_api}?apikey={API_KEY}&formato=json" target="_blank" style="color:{C["primary"]};font-weight:600;word-break:break-all;">
                  /v1/tasas/{tipo_api}?apikey=...&formato=json
                </a><br><br>
                Si el endpoint requiere un parámetro <code>periodo</code> (ej: <code>202412</code>), la estructura de tasas de la CMF puede haber cambiado. Consulta la documentación en <a href="https://api.cmfchile.cl" target="_blank" style="color:{C["primary"]};">api.cmfchile.cl</a>
              </div>
            </div>""", unsafe_allow_html=True)
        else:
            campo_tasa, campo_banco = _detectar_campos(tasas_actuales)
            def _to_float(v):
                try: return float(str(v).replace(",","."))
                except: return None
            rows_validos = [{**r, "_tasa_num": _to_float(r.get(campo_tasa))} for r in tasas_actuales if _to_float(r.get(campo_tasa)) is not None]
            rows_validos.sort(key=lambda x: x["_tasa_num"])

            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:16px 0 10px;">Ranking · {tipo_label}</div>', unsafe_allow_html=True)
            medals=["🥇","🥈","🥉"]; min_t=rows_validos[0]["_tasa_num"] if rows_validos else 1; max_t=rows_validos[-1]["_tasa_num"] if rows_validos else 1; rango=max_t-min_t or 1
            html_rank='<div style="display:flex;flex-direction:column;gap:7px;margin-bottom:4px;">'
            for i,row in enumerate(rows_validos):
                t=row["_tasa_num"]; banco=row.get(campo_banco,"—"); pct=(t-min_t)/rango*100; bc=BANK_COLOR.get(banco,C["primary"])
                med=medals[i] if i<3 else f'<span style="font-size:11px;color:{C["n400"]};min-width:22px;display:inline-block;text-align:center;">#{i+1}</span>'
                html_rank+=f'<div style="display:flex;align-items:center;gap:10px;"><span style="font-size:15px;min-width:22px;text-align:center;">{med}</span><div style="flex:1;"><div style="display:flex;justify-content:space-between;margin-bottom:3px;"><span style="font-size:12px;font-weight:600;color:{C["n900"]};">{banco}</span><span style="font-size:12px;font-weight:700;color:{bc};">{t:.2f}%</span></div><div style="background:{C["n100"]};border-radius:99px;height:5px;"><div style="background:{bc};width:{max(4,pct):.1f}%;height:5px;border-radius:99px;"></div></div></div></div>'
            st.markdown(html_rank+'</div>', unsafe_allow_html=True)

            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:24px 0 10px;">Simulador de cuota</div>', unsafe_allow_html=True)
            with st.container():
                col_m,col_p=st.columns(2)
                with col_m:
                    st.markdown('<span style="font-size:12px;font-weight:600;color:#64748b;">Monto ($)</span>', unsafe_allow_html=True)
                    monto_sim=st.number_input("Monto",min_value=100_000,max_value=500_000_000,value=5_000_000,step=500_000,label_visibility="collapsed",key="sim_monto")
                with col_p:
                    st.markdown('<span style="font-size:12px;font-weight:600;color:#64748b;">Plazo (meses)</span>', unsafe_allow_html=True)
                    plazo_sim=st.number_input("Plazo",min_value=6,max_value=360,value=36,step=6,label_visibility="collapsed",key="sim_plazo")

            fmt_clp=lambda n: f"${n:,.0f}".replace(",","."); sim_cards=""; mejor_cuota=None
            for row in rows_validos[:8]:
                banco=row.get(campo_banco,"—"); tasa=row["_tasa_num"]; sim=simular_cuota(monto_sim,tasa,plazo_sim)
                if not sim: continue
                if mejor_cuota is None: mejor_cuota=sim["cuota"]
                bc=BANK_COLOR.get(banco,C["primary"]); is_best=sim["cuota"]==mejor_cuota
                ahorro="" if is_best else f'<div style="font-size:10px;color:{C["n400"]};margin-top:2px;">+{fmt_clp(sim["cuota"]-mejor_cuota)}/mes vs mejor</div>'
                sim_cards+=f'<div class="ratio-card {"leader" if is_best else ""}"><div style="display:flex;align-items:center;gap:8px;margin-bottom:2px;"><div style="width:3px;height:34px;border-radius:2px;background:{bc};flex-shrink:0;"></div><div><div class="ratio-meta">tasa anual</div><div style="display:flex;align-items:center;gap:5px;"><span class="ratio-banco">{banco}</span>{"<span class=\'badge-lider\'>MEJOR</span>" if is_best else ""}</div></div></div><div class="ratio-divider"></div><div style="font-size:10px;color:{C["n400"]};margin-bottom:2px;">{tasa:.2f}% anual</div><span class="ratio-num" style="color:{bc};font-size:22px;">{fmt_clp(sim["cuota"])}</span><div style="font-size:10px;color:{C["n600"]};margin-top:2px;">/mes · {plazo_sim} meses</div>{ahorro}<div style="font-size:10px;color:{C["n400"]};margin-top:6px;">Total interés: {fmt_clp(sim["interes"])} ({sim["costo_pct"]}%)</div></div>'
            st.markdown('<div class="ratio-row">'+sim_cards+'</div>', unsafe_allow_html=True)

            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C["primary"]};text-transform:uppercase;letter-spacing:1.2px;margin:24px 0 10px;">Evolución histórica · 12 meses</div>', unsafe_allow_html=True)
            with st.spinner("Cargando histórico..."): hist=get_tasas_historicas_cmf(tipo_api,12)
            if hist:
                campo_t2,_=_detectar_campos(hist); periodos_data={}
                for r in hist:
                    p=r.get("_periodo",""); val=_to_float(r.get(campo_t2))
                    if p and val: periodos_data.setdefault(p,[]).append(val)
                periodos_ord=sorted(periodos_data.keys()); promedios=[sum(periodos_data[p])/len(periodos_data[p]) for p in periodos_ord]; labels=[f"{p[:4]}-{p[4:]}" for p in periodos_ord]
                fig_hist=go.Figure()
                fig_hist.add_trace(go.Scatter(x=labels,y=promedios,fill="tozeroy",fillcolor="rgba(77,107,255,0.07)",line=dict(color=C["primary"],width=2.5),mode="lines+markers",marker=dict(color=C["primary"],size=6),name="Tasa promedio"))
                if tpm_data:
                    tpm_dict={t["fecha"][:7].replace("-",""):t["tpm"] for t in tpm_data}; tpm_vals=[tpm_dict.get(p) for p in periodos_ord]
                    if any(v is not None for v in tpm_vals): fig_hist.add_trace(go.Scatter(x=labels,y=tpm_vals,line=dict(color="#06d6a0",width=2,dash="dot"),mode="lines",name="TPM BCCH"))
                fig_hist.update_layout(template="plotly_white",height=260,margin=dict(l=0,r=0,t=10,b=10),legend=dict(orientation="h",y=-0.35,font=dict(family="Plus Jakarta Sans",size=11)),yaxis=dict(ticksuffix="%"))
                st.plotly_chart(fig_hist,use_container_width=True)
            else:
                st.markdown(f'<div style="font-size:12px;color:{C["n400"]};text-align:center;padding:20px;">Sin datos históricos disponibles.</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
#  EMPRESA
# ══════════════════════════════════════════════════════
with tab_empresa:
    st.info(f"Módulo de Empresas conectado a: {RUTA_BD}")

# ══════════════════════════════════════════════════════
#  MI EMPRESA
# ══════════════════════════════════════════════════════
with tab_mi_empresa:
    render_mi_empresa(C, BANK_COLOR)

# ══════════════════════════════════════════════════════
#  CONFIGURACION
# ══════════════════════════════════════════════════════
with tab_config:
    st.markdown(f'<span style="font-size:15px;color:{C["n600"]};">Configuracion — próximamente.</span>', unsafe_allow_html=True)